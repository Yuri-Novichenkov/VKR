"""
Бенчмарк оптимизаций:
  1. fast_knn=True  — torch_cluster (GPU-only, пропускается если не установлен)
  2. INT8            — dynamic quantization (CPU)

Для каждой модели измеряем:
  - Время инференса FP32 на GPU (baseline, N=4096)
  - Время инференса FP32 + fast_knn на GPU (N=4096)
  - Время инференса FP16 + fast_knn на GPU (N=4096)
  - Время инференса INT8 на CPU (N=1024, т.к. O(N²) kNN нереален при N=4096)
  - Время инференса FP32 на CPU (N=1024, baseline для сравнения с INT8)

Примечание: CPU-бенчмарк использует N=1024 вместо N=4096, поскольку
матричный kNN O(N²) делает граф-модели неприемлемо медленными при N=4096.
Ускорение INT8 считается относительно FP32 CPU при том же N=1024.

Запуск:
    python scripts/optimize_benchmark.py
"""

import sys, warnings, time, torch, numpy as np
from pathlib import Path

# Подавляем DeprecationWarning для torch.ao.quantization (deprecated в 2.10)
warnings.filterwarnings(
    "ignore",
    message="torch.ao.quantization is deprecated",
    category=DeprecationWarning,
)

sys.path.insert(0, ".")
from src.models import build_model

# Проверяем доступность torch_cluster (GPU-only оптимизация)
try:
    import torch_cluster  # noqa: F401
    HAS_TORCH_CLUSTER = True
except ImportError:
    HAS_TORCH_CLUSTER = False
    print("INFO: torch_cluster не установлен — fast_knn (GPU) будет пропущен.")
    print("      На Windows с PyTorch 2.10+: PyG wheels для cu130 ещё не выпущены.")
    print()

# ── Все 8 моделей ─────────────────────────────────────────────────────────────
CONFIGS = [
    ("PointNet",         "pointnet",         "checkpoints/loss_sweep/pointnet/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointNet++",       "pointnet++",       "checkpoints/loss_sweep/pointnet++/segmentation/loss_lovasz_g2p0__cb_effective_b0p99999/mar16/best_model.pth"),
    ("DGCNN",            "dgcnn",            "checkpoints/loss_sweep/dgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN",           "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN-GATv2",     "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_gatv2_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN-LocalWin",  "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_local_window_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNNFlash",      "ldgcnn_flash",     "checkpoints/loss_sweep/ldgcnn_flash/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointTransformer", "pointtransformer", "checkpoints/pointtransformer/segmentation/cb_effective_b0p999/mar16/best_model.pth"),
]

DEFAULTS = dict(k=20, k_small=20, k_large=40, attention_type="none",
                attention_k=16, attention_heads=4, attention_dropout=0.1, pt_k=16)

GPU_N   = 4096   # N для GPU-бенчмарка
CPU_N   = 1024   # N для CPU-бенчмарка (O(N²) kNN иначе зависает)
WARMUP  = 5
REPS_GPU = 30
REPS_CPU = 5     # меньше повторений — граф-модели на CPU медленные даже при N=1024


# ── Загрузка модели ───────────────────────────────────────────────────────────

def load_model(path, fast_knn=False):
    ck = torch.load(path, map_location="cpu", weights_only=False)
    meta = {**DEFAULTS}
    for key in DEFAULTS:
        if ck.get(key) is not None:
            meta[key] = ck[key]
    mtype = ck.get("model_type", "pointnet")
    m = build_model(
        mtype, task="segmentation",
        num_classes=ck["num_classes"], num_features=ck["num_features"],
        k=meta["k"], k_small=meta["k_small"], k_large=meta["k_large"],
        attention_type=meta["attention_type"], attention_k=meta["attention_k"],
        attention_heads=meta["attention_heads"], attention_dropout=meta["attention_dropout"],
        pt_k=meta["pt_k"],
        fast_knn=fast_knn,
    )
    m.load_state_dict(ck["model_state_dict"], strict=False)
    return m.eval(), ck["num_features"]


def _forward(model, x):
    """Запускает модель, распаковывает кортеж (PointNet возвращает (x, t1, t2))."""
    out = model(x)
    return out[0] if isinstance(out, tuple) else out


# ── Бенчмарки ─────────────────────────────────────────────────────────────────

def bench_gpu(model, num_features, device, dtype=torch.float32):
    """Среднее время инференса (мс) на GPU через CUDA events."""
    model = model.to(device)
    if dtype == torch.float16:
        model = model.half()
    x = torch.randn(1, GPU_N, num_features, device=device, dtype=dtype)
    with torch.no_grad():
        for _ in range(WARMUP):
            _forward(model, x)
    torch.cuda.synchronize()
    s = torch.cuda.Event(enable_timing=True)
    e = torch.cuda.Event(enable_timing=True)
    times = []
    with torch.no_grad():
        for _ in range(REPS_GPU):
            s.record(); _forward(model, x); e.record()
            torch.cuda.synchronize()
            times.append(s.elapsed_time(e))
    model.cpu()
    torch.cuda.empty_cache()
    return float(np.mean(times))


def bench_cpu(model, num_features):
    """Среднее время инференса на CPU (мс), N=CPU_N=1024."""
    x = torch.randn(1, CPU_N, num_features)
    with torch.no_grad():
        for _ in range(WARMUP):
            _forward(model, x)
    times = []
    with torch.no_grad():
        for _ in range(REPS_CPU):
            t0 = time.perf_counter()
            _forward(model, x)
            times.append((time.perf_counter() - t0) * 1000)
    return float(np.mean(times))


def apply_int8(model):
    """Dynamic INT8 quantization: квантует Linear, Conv1d и Conv2d.
    Conv2d критичен для EdgeConv-based моделей (DGCNN/LDGCNN/LDGCNNFlash).
    """
    return torch.quantization.quantize_dynamic(
        model,
        {torch.nn.Linear, torch.nn.Conv1d, torch.nn.Conv2d},
        dtype=torch.qint8,
    )


# ── Основной цикл ─────────────────────────────────────────────────────────────

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"GPU-бенчмарк: {device}  N={GPU_N}")
print(f"CPU-бенчмарк: cpu  N={CPU_N}  (N<4096 чтобы избежать зависания O(N^2) kNN)\n")

results = []

for label, mtype, path in CONFIGS:
    if not Path(path).exists():
        print(f"WARN: чекпоинт не найден, пропускаю: {path}")
        continue

    print(f"\n{'='*52}")
    print(f"  {label}")
    print(f"{'='*52}")

    # ── FP32 GPU baseline ──────────────────────────────────────────────────
    model, num_feat = load_model(path, fast_knn=False)
    if device.type == "cuda":
        t_fp32 = bench_gpu(model, num_feat, device, torch.float32)
        print(f"  FP32 GPU (N={GPU_N}):      {t_fp32:.2f} ms")
    else:
        t_fp32 = float("nan")
        print(f"  FP32 GPU: пропущено (нет CUDA)")

    # ── FP32 + fast_knn GPU ────────────────────────────────────────────────
    t_fp32_fk = float("nan")
    t_fp16_fk = float("nan")
    if HAS_TORCH_CLUSTER and device.type == "cuda":
        model_fk, _ = load_model(path, fast_knn=True)
        t_fp32_fk = bench_gpu(model_fk, num_feat, device, torch.float32)
        print(f"  FP32+fast_knn GPU:    {t_fp32_fk:.2f} ms  (x{t_fp32/t_fp32_fk:.2f})")
        model_fk16, _ = load_model(path, fast_knn=True)
        t_fp16_fk = bench_gpu(model_fk16, num_feat, device, torch.float16)
        print(f"  FP16+fast_knn GPU:    {t_fp16_fk:.2f} ms  (x{t_fp32/t_fp16_fk:.2f})")
    else:
        print(f"  fast_knn: пропущено")

    # ── INT8 CPU (N=1024) ──────────────────────────────────────────────────
    model_int8, _ = load_model(path, fast_knn=False)
    model_int8 = apply_int8(model_int8)
    t_int8 = bench_cpu(model_int8, num_feat)

    model_fp32_cpu, _ = load_model(path, fast_knn=False)
    t_fp32_cpu = bench_cpu(model_fp32_cpu, num_feat)

    int8_speedup = t_fp32_cpu / t_int8
    print(f"  FP32 CPU (N={CPU_N}):      {t_fp32_cpu:.1f} ms")
    print(f"  INT8 CPU (N={CPU_N}):      {t_int8:.1f} ms  (x{int8_speedup:.2f} vs FP32 CPU)")

    def _x(a, b):
        if isinstance(b, float) and (b != b or b == 0):
            return None
        return round(a / b, 2) if b else None

    results.append({
        "model":        label,
        "fp32_gpu_ms":  round(t_fp32, 2) if t_fp32 == t_fp32 else "—",
        "fp32_fk_ms":   round(t_fp32_fk, 2) if HAS_TORCH_CLUSTER else "—",
        "fp32_fk_x":    _x(t_fp32, t_fp32_fk) if HAS_TORCH_CLUSTER else "—",
        "fp16_fk_ms":   round(t_fp16_fk, 2) if HAS_TORCH_CLUSTER else "—",
        "fp16_fk_x":    _x(t_fp32, t_fp16_fk) if HAS_TORCH_CLUSTER else "—",
        "fp32_cpu_ms":  round(t_fp32_cpu, 1),
        "int8_cpu_ms":  round(t_int8, 1),
        "int8_x":       round(int8_speedup, 2),
        "gpu_N":        GPU_N,
        "cpu_N":        CPU_N,
    })


# ── Консольная таблица ────────────────────────────────────────────────────────

def _fmt(val, fmt):
    if val is None or val == "—":
        return "—"
    try:
        if isinstance(val, float) and val != val:
            return "—"
        return format(val, fmt)
    except (TypeError, ValueError):
        return str(val)

print("\n\n" + "="*80)
print(f"ИТОГО  |  GPU N={GPU_N}  |  CPU N={CPU_N}")
print("="*80)
print(f"{'Модель':<16} {'FP32 GPU':>9} {'FP32+fkNN':>10} {'x':>5} {'FP16+fkNN':>10} {'x':>5} {'FP32 CPU':>9} {'INT8 CPU':>9} {'x':>5}")
print("-"*80)
for r in results:
    print(
        f"{r['model']:<16}"
        f" {_fmt(r['fp32_gpu_ms'],'.2f'):>9}"
        f" {_fmt(r['fp32_fk_ms'],'.2f'):>10}"
        f" {_fmt(r['fp32_fk_x'],'.2f'):>5}"
        f" {_fmt(r['fp16_fk_ms'],'.2f'):>10}"
        f" {_fmt(r['fp16_fk_x'],'.2f'):>5}"
        f" {_fmt(r['fp32_cpu_ms'],'.1f'):>9}"
        f" {_fmt(r['int8_cpu_ms'],'.1f'):>9}"
        f" {_fmt(r['int8_x'],'.2f'):>5}"
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws.title = "INT8 оптимизация"

HDR   = PatternFill("solid", start_color="BDD7EE")
GREEN = PatternFill("solid", start_color="C6EFCE")
thin  = Side(style="thin", color="AAAAAA")
brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
PALETTE = ["DEEAF1","E2EFDA","FFF2CC","FCE4D6","EAE1F4","D9EAD3","FDE9D9","E8F5E9"]
mc = {r["model"]: PALETTE[i % len(PALETTE)] for i, r in enumerate(results)}

headers = [
    f"Модель",
    f"FP32 GPU\n(мс, N={GPU_N})",
    f"FP32+fastKNN GPU\n(мс, N={GPU_N})",
    "Уск. kNN (x)",
    f"FP16+fastKNN GPU\n(мс, N={GPU_N})",
    "Уск. FP16 (x)",
    f"FP32 CPU\n(мс, N={CPU_N})",
    f"INT8 CPU\n(мс, N={CPU_N})",
    "Уск. INT8 (x)",
]
ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col)
    c.value = h
    c.font = Font(bold=True, name="Arial", size=9)
    c.fill = HDR
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd

for r_idx, r in enumerate(results, 2):
    fill = PatternFill("solid", start_color=mc[r["model"]])
    vals = [
        r["model"], r["fp32_gpu_ms"], r["fp32_fk_ms"], r["fp32_fk_x"],
        r["fp16_fk_ms"], r["fp16_fk_x"], r["fp32_cpu_ms"], r["int8_cpu_ms"], r["int8_x"],
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(r_idx, col)
        c.value = val if val != "—" else "—"
        c.font = Font(name="Arial", size=10)
        c.fill = fill
        c.border = brd
        c.alignment = Alignment(horizontal="center")
    # Подсветить INT8 speedup зелёным если > 1.3
    if isinstance(r["int8_x"], (int, float)) and r["int8_x"] >= 1.3:
        ws.cell(r_idx, 9).fill = GREEN
        ws.cell(r_idx, 9).font = Font(bold=True, name="Arial", size=10)

col_widths = [16, 13, 18, 12, 18, 13, 13, 13, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w
for r in range(1, len(results) + 2):
    ws.row_dimensions[r].height = 36

# График: INT8 speedup по моделям
ws2 = wb.create_sheet("Графики")
ws2.append(["Модель", f"FP32 CPU (N={CPU_N})", f"INT8 CPU (N={CPU_N})"])
for r in results:
    ws2.append([r["model"], r["fp32_cpu_ms"], r["int8_cpu_ms"]])

chart = BarChart()
chart.type = "col"
chart.title = f"FP32 vs INT8 инференс на CPU (N={CPU_N})"
chart.y_axis.title = "мс"
chart.style = 10
chart.width = 22
chart.height = 14
data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=len(results) + 1)
cats = Reference(ws2, min_col=1, min_row=2, max_row=len(results) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, "A" + str(len(results) + 4))

# Заметка о N
ws.cell(len(results) + 3, 1).value = (
    f"* GPU-бенчмарк: N={GPU_N} точек. CPU-бенчмарк: N={CPU_N} точек "
    f"(O(N^2) kNN при N={GPU_N} зависает на CPU для граф-моделей)."
)
ws.cell(len(results) + 3, 1).font = Font(italic=True, name="Arial", size=9)

out = "results/optimization_results.xlsx"
Path("results").mkdir(exist_ok=True)
wb.save(out)
print(f"\nСохранено: {out}")
