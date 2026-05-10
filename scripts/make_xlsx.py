from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws.title = "Архитектура и время"

HDR_FILL  = PatternFill("solid", start_color="BDD7EE")
ORANGE    = PatternFill("solid", start_color="F4B942")
YELLOW    = PatternFill("solid", start_color="FFEB9C")
TOTAL_FILL = PatternFill("solid", start_color="D9D9D9")
MODEL_COLORS = ["DEEAF1","E2EFDA","FFF2CC","FCE4D6","EAE1F4","D9EAD3","FDE9D9","E8F0FE"]

thin = Side(style="thin", color="AAAAAA")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Модель", "Блок", "Операция", "Вход → Выход", "мс", "%"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col)
    c.font = Font(bold=True, name="Arial", size=10)
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd
ws.freeze_panes = "A2"

rows = [
    ("PointNet","input_transform","T-Net 3x3: выравнивание XYZ","3->64->1024->3x3",1.81,19.8),
    ("PointNet","conv1-conv2","Shared MLP по точкам","9->64->64",0.52,5.7),
    ("PointNet","feature_transform","T-Net 64x64: выравнивание фич","64->64->1024->64x64",2.17,23.8),
    ("PointNet","conv3-conv4","Shared MLP","64->128->1024",1.04,11.4),
    ("PointNet","Global max pool","Агрегация -> глобальный вектор","1024xN->1024",None,None),
    ("PointNet","conv5-conv8","MLP-сегментатор (1088 = 64+1024)","1088->512->256->128->11",2.29,25.1),
    ("PointNet","ИТОГО","","",9.1,None),

    ("PointNet++","sa1","FPS 4096->512, Ball query r=0.2, nsample=32, MLP","9->64->64->128",378.9,77.8),
    ("PointNet++","sa2","FPS 512->128, Ball query r=0.4, nsample=64, MLP","131->128->256",100.3,20.6),
    ("PointNet++","sa3","Глобальный Set Abstraction, MLP","259->256->512->1024",1.0,0.2),
    ("PointNet++","fp3-fp1","IDW-интерполяция (декодер U-Net)","1280->256, 384->128, 134->128",6.3,1.3),
    ("PointNet++","classifier","Conv1d + Dropout","128->128->11",0.5,0.1),
    ("PointNet++","ИТОГО","","",487.0,None),

    ("DGCNN","ec1","kNN(XYZ, k=20) + EdgeConv","18->64",1.5,8.3),
    ("DGCNN","ec2","kNN(feat64, k=20) + EdgeConv","128->64",1.7,9.4),
    ("DGCNN","ec3","kNN(feat64, k=20) + EdgeConv","128->128",3.3,18.0),
    ("DGCNN","ec4","kNN(feat128, k=20) + EdgeConv","256->256",11.2,60.2),
    ("DGCNN","conv1-conv3","MLP-классификатор (concat=512)","512->256->128->11",0.7,4.0),
    ("DGCNN","ИТОГО","","",18.5,None),

    ("LDGCNN","ec1","Multi-scale kNN(XYZ, k=20/40) + EdgeConv","18->64",4.9,5.5),
    ("LDGCNN","ec2","Multi-scale kNN(feat, k=20/40) + EdgeConv","128->64",15.9,17.8),
    ("LDGCNN","ec3","Multi-scale kNN(feat, k=20/40) + EdgeConv","128->128",20.7,23.1),
    ("LDGCNN","ec4","Multi-scale kNN(feat, k=20/40) + EdgeConv","256->256",47.1,52.6),
    ("LDGCNN","conv1-conv3","MLP-классификатор","512->256->128->11",0.8,0.9),
    ("LDGCNN","ИТОГО","","",89.5,None),

    ("LDGCNN GAT","ec1-ec2","Multi-scale EdgeConv","18->64, 128->64",22.9,22.7),
    ("LDGCNN GAT","attention","GATv2: kNN(k=16) -> веса внимания -> взвеш. сумма","64, heads=4",1.7,1.7),
    ("LDGCNN GAT","ec3-ec4","Multi-scale EdgeConv","128->128, 256->256",75.2,74.6),
    ("LDGCNN GAT","conv1-conv3","MLP-классификатор","512->256->128->11",0.8,0.8),
    ("LDGCNN GAT","ИТОГО","","",100.8,None),

    ("LDGCNN WINDOW","ec1-ec2","Multi-scale EdgeConv","18->64, 128->64",23.7,23.0),
    ("LDGCNN WINDOW","attention","Local Window: sliding window self-attn в kNN(k=16)","64, heads=4",1.2,1.2),
    ("LDGCNN WINDOW","ec3-ec4","Multi-scale EdgeConv","128->128, 256->256",77.1,74.7),
    ("LDGCNN WINDOW","conv1-conv3","MLP-классификатор","512->256->128->11",1.0,1.0),
    ("LDGCNN WINDOW","ИТОГО","","",103.2,None),

    ("LDGCNNFlash","ec1","kNN(XYZ, k=20) + EdgeConv","9->64",2.1,5.8),
    ("LDGCNNFlash","attn1","Flash Self-Attention по N=4096 точкам","64, heads=4",2.5,6.9),
    ("LDGCNNFlash","ec2","Linked EdgeConv: concat(x,a1)->граф","73->128",6.3,17.5),
    ("LDGCNNFlash","attn2","Flash Self-Attention","128, heads=4",3.8,10.6),
    ("LDGCNNFlash","ec3","Linked EdgeConv: concat(x,a1,a2)->граф","201->256",10.4,28.9),
    ("LDGCNNFlash","attn3","Flash Self-Attention (матрица 4096x4096)","256, heads=4",6.1,16.9),
    ("LDGCNNFlash","head","MLP: concat(a1,a2,a3)","448->256->11",0.8,2.2),
    ("LDGCNNFlash","ИТОГО","","",36.0,None),

    ("PointTransformer","stem","Linear-проекция","9->32",0.3,0.1),
    ("PointTransformer","enc0","PT Block: векторное внимание kNN(k=16)","32, N=4096",1.7,0.5),
    ("PointTransformer","td1","FPS 4096->1024 + kNN + агрегация","32->64",254.7,69.1),
    ("PointTransformer","enc1","PT Block","64, N=1024",2.0,0.5),
    ("PointTransformer","td2","FPS 1024->256 + kNN + агрегация","64->128",65.5,17.8),
    ("PointTransformer","enc2-enc4","PT Blocks на малых облаках","128/256/512",5.5,1.5),
    ("PointTransformer","td3-td4","FPS 256->64, 64->16","128->256, 256->512",24.1,6.5),
    ("PointTransformer","tu0-tu3 + dec0-dec3","TransitionUp + PT Blocks (декодер)","512->256->128->64->32",14.7,4.0),
    ("PointTransformer","head","Linear","32->11",0.2,0.1),
    ("PointTransformer","ИТОГО","","",368.5,None),
]

model_list = ["PointNet","PointNet++","DGCNN","LDGCNN","LDGCNN GAT","LDGCNN WINDOW","LDGCNNFlash","PointTransformer"]
model_color = {m: MODEL_COLORS[i] for i, m in enumerate(model_list)}

for r_idx, row in enumerate(rows, 2):
    model, block, op, dims, ms, pct = row
    is_total = block == "ИТОГО"
    fill = TOTAL_FILL if is_total else PatternFill("solid", start_color=model_color[model])

    for c_idx, val in enumerate([model, block, op, dims, ms, pct], 1):
        cell = ws.cell(r_idx, c_idx)
        cell.value = "—" if val is None else val
        cell.font = Font(bold=is_total, name="Arial", size=10)
        cell.fill = fill
        cell.border = brd
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if c_idx in (5, 6) and isinstance(val, float):
            cell.number_format = "0.0"

    if pct is not None:
        pct_cell = ws.cell(r_idx, 6)
        if pct >= 50:
            pct_cell.fill = ORANGE
            pct_cell.font = Font(bold=True, name="Arial", size=10)
        elif pct >= 15:
            pct_cell.fill = YELLOW

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 46
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 8
for r_idx in range(2, len(rows) + 2):
    ws.row_dimensions[r_idx].height = 32

# Sheet 2: bar chart
ws2 = wb.create_sheet("Время инференса")
ws2.append(["Модель", "мс (итого)"])
totals = [
    ("PointNet", 9.1), ("PointNet++", 487.0), ("DGCNN", 18.5),
    ("LDGCNN", 89.5), ("LDGCNN GAT", 100.8), ("LDGCNN WINDOW", 103.2),
    ("LDGCNNFlash", 36.0), ("PointTransformer", 368.5),
]
for row in totals:
    ws2.append(list(row))
ws2["A1"].font = Font(bold=True, name="Arial")
ws2["B1"].font = Font(bold=True, name="Arial")
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14

chart = BarChart()
chart.type = "col"
chart.title = "Время инференса по моделям (мс, batch=1, N=4096)"
chart.y_axis.title = "мс"
chart.x_axis.title = "Модель"
chart.style = 10
chart.width = 22
chart.height = 14
data_ref = Reference(ws2, min_col=2, min_row=1, max_row=9)
cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=9)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws2.add_chart(chart, "D2")

out = r"D:\Уник\VKR\results\model_architecture_timing.xlsx"
wb.save(out)
print("saved:", out)
