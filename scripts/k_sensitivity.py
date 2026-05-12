"""
Анализ чувствительности к числу ближайших соседей k.

Для каждой модели и каждого значения k измеряет:
  1. Скорость инференса (мс, GPU)
  2. Точность (mIoU) на тестовом наборе Mar16 — с scene-level voting

Модели: DGCNN, LDGCNN, LDGCNNFlash, PointTransformer
Значения k: 5, 8, 10, 16, 20, 24, 32

Запуск:
    # только скорость (быстро, ~2 мин):
    python scripts/k_sensitivity.py --speed_only

    # скорость + точность (долго, ~60-90 мин):
    python scripts/k_sensitivity.py
"""

import sys, argparse, torch, numpy as np
from pathlib import Path

sys.path.insert(0, ".")
from src.models import build_model
from src.utils.metrics import calculate_metrics

CONFIGS = [
    ("DGCNN",         "dgcnn",         "checkpoints/loss_sweep/dgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN",        "ldgcnn",        "checkpoints/loss_sweep/ldgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNNFlash",   "ldgcnn_flash",  "checkpoints/loss_sweep/ldgcnn_flash/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointTransformer", "pointtransformer", "checkpoints/pointtransformer/segmentation/cb_effective_b0p999/mar16/best_model.pth"),
]
DEFAULTS = dict(k=20, k_small=20, k_large=40, attention_type="none",
                attention_k=16, attention_heads=4, attention_dropout=0.1, pt_k=16)

K_VALUES = [5, 8, 10, 16, 20, 24, 32]
# Для PointTransformer базовый k=16, тестируем другой диапазон
K_VALUES_PT = [4, 8, 12, 16, 20, 24, 32]

N_POINTS = 4096
WARMUP, REPS = 5, 20
TEST_DATA = "Files/Mar16/LiDAR/Mar16_test_GroundTruth.laz"


# ── Патчинг k в модели без перезагрузки ───────────────────────────────────────

def set_k(model, mtype, k):
    """Изменяет k во всех слоях модели. Для LDGCNN k_large = k*2."""
    if mtype == "dgcnn":
        model.k = k

    elif mtype == "ldgcnn":
        k_large = k * 2
        for ec in [model.ec1, model.ec2, model.ec3, model.ec4]:
            ec.k_small = k
            ec.k_large = k_large
        model.attention_k = k  # для attention-вариантов

    elif mtype == "ldgcnn_flash":
        for ec in [model.ec1, model.ec2, model.ec3]:
            ec.k = k

    elif mtype == "pointtransformer":
        for td in [model.td1, model.td2, model.td3, model.td4]:
            td.k = k
        # PT-блоки используют k из PointTransformerLayer
        for block in [model.enc0, model.enc1, model.enc2, model.enc3, model.enc4,
                      model.dec0, model.dec1, model.dec2, model.dec3]:
            block.k     = k
            block.pt.k  = k


def get_baseline_k(mtype, ck):
    """Возвращает базовый k из чекпоинта."""
    if mtype == "dgcnn":
        return ck.get("k", 20)
    elif mtype == "ldgcnn":
        return ck.get("k_small", 20)
    elif mtype == "ldgcnn_flash":
        return ck.get("k", 20)
    elif mtype == "pointtransformer":
        return ck.get("pt_k", 16)
    return 20


# ── Загрузка модели ───────────────────────────────────────────────────────────

def load_model(path):
    ck = torch.load(path, map_location="cpu", weights_only=False)
    meta = {**DEFAULTS}
    for key in DEFAULTS:
        if ck.get(key) is not None:
            meta[key] = ck[key]
    mtype = ck.get("model_type", "pointnet")
    m = build_model(
        mtype, task="segmentation",
        num_classes=ck["num_classes"], num_features=ck["num_features"],
        k=meta["k"], k_small=meta["k_small"], k_large=meta["k_large"],
        attention_type=meta["attention_type"], attention_k=meta["attention_k"],
        attention_heads=meta["attention_heads"], attention_dropout=meta["attention_dropout"],
        pt_k=meta["pt_k"],
    )
    m.load_state_dict(ck["model_state_dict"], strict=False)
    return m.eval(), ck["num_features"], ck["num_classes"], mtype, ck


# ── Бенчмарк скорости ─────────────────────────────────────────────────────────

def bench_speed(model, num_features, device):
    model = model.to(device)
    x = torch.randn(1, N_POINTS, num_features, device=device)
    with torch.no_grad():
        for _ in range(WARMUP):
            model(x)
    torch.cuda.synchronize()
    s = torch.cuda.Event(enable_timing=True)
    e = torch.cuda.Event(enable_timing=True)
    times = []
    with torch.no_grad():
        for _ in range(REPS):
            s.record(); model(x); e.record()
            torch.cuda.synchronize()
            times.append(s.elapsed_time(e))
    model.cpu()
    torch.cuda.empty_cache()
    return float(np.mean(times))


# ── Создание тестового датасета ───────────────────────────────────────────────

def make_test_dataset():
    """Создаёт датасет один раз с кэшированием на диск."""
    from src.data.dataset import LiDARDataset
    return LiDARDataset(
        TEST_DATA, num_points=N_POINTS, augment=False,
        task="segmentation",
        cache_dir="cache", cache_mode="write",
    )


# ── Оценка точности (mIoU с voting) ──────────────────────────────────────────

def eval_accuracy(model, num_classes, device, dataset):
    """Оценивает mIoU с scene-level voting. Датасет передаётся снаружи."""
    from torch.utils.data import DataLoader

    loader = DataLoader(dataset, batch_size=4, shuffle=False,
                        num_workers=2, pin_memory=True)

    num_total = len(dataset.features)
    vote_counts = np.zeros((num_total, num_classes), dtype=np.int32)
    model = model.to(device)

    with torch.no_grad():
        sample_offset = 0
        for features, labels in loader:
            features = features.to(device)
            logits = model(features)                          # (B, N, C)
            preds = logits.argmax(dim=-1).cpu().numpy()       # (B, N)
            bsz = preds.shape[0]
            for b in range(bsz):
                idx = dataset.get_cloud_point_indices(sample_offset + b)
                np.add.at(vote_counts, (idx, preds[b]), 1)
            sample_offset += bsz

    aggregated = vote_counts.argmax(axis=1)
    true_labels = np.asarray(dataset.labels, dtype=np.int64)
    metrics = calculate_metrics(
        torch.from_numpy(aggregated),
        torch.from_numpy(true_labels),
        num_classes=num_classes,
    )
    model.cpu()
    torch.cuda.empty_cache()
    return float(metrics["mean_iou"]) * 100


# ── Основной цикл ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--speed_only", action="store_true",
                        help="Только скорость, без оценки точности")
    parser.add_argument("--models", nargs="+", default=None,
                        help="Запустить только для указанных моделей, напр. DGCNN LDGCNNFlash")
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    test_exists = Path(TEST_DATA).exists()
    if not args.speed_only and not test_exists:
        print(f"WARN: тестовый файл не найден ({TEST_DATA}), запускаю только скорость.")
        args.speed_only = True

    # Датасет создаётся один раз и переиспользуется для всех моделей и k.
    # cache_mode="write" пишет кэш при первом запуске; последующие запуски читают с диска.
    test_dataset = None
    if not args.speed_only:
        print(f"Загрузка тестового датасета (первый раз может занять минуту)...")
        test_dataset = make_test_dataset()
        print(f"  Датасет: {len(test_dataset)} окон из {len(test_dataset.features)} точек")

    all_results = []  # list of dicts

    for label, mtype, path in CONFIGS:
        if args.models and label not in args.models:
            continue

        print(f"\n{'='*55}")
        print(f"  {label}  (базовый чекпоинт: {Path(path).parent.parent.name})")
        print(f"{'='*55}")

        model, num_feat, num_cls, mtype_ck, ck = load_model(path)
        baseline_k = get_baseline_k(mtype_ck, ck)
        k_list = K_VALUES_PT if mtype == "pointtransformer" else K_VALUES

        # Предварительно измеряем baseline для расчёта speedup
        set_k(model, mtype_ck, baseline_k)
        baseline_ms = bench_speed(model, num_feat, device)
        print(f"  baseline k={baseline_k}: {baseline_ms:.2f} ms")
        print(f"  {'k':>4}  {'мс':>8}  {'speedup':>8}  {'mIoU%':>8}")
        print(f"  {'-'*38}")

        for k in k_list:
            set_k(model, mtype_ck, k)
            ms = bench_speed(model, num_feat, device)

            miou = None
            if not args.speed_only:
                miou = eval_accuracy(model, num_cls, device, test_dataset)

            speedup = f"{baseline_ms/ms:.2f}x"
            miou_str = f"{miou:.2f}" if miou is not None else "—"
            marker = " ← baseline" if k == baseline_k else ""
            print(f"  {k:>4}  {ms:>8.2f}  {speedup:>8}  {miou_str:>8}{marker}")

            all_results.append({
                "model": label, "k": k, "baseline_k": baseline_k,
                "ms": round(ms, 2),
                "speedup": round(baseline_ms / ms, 2),
                "miou": round(miou, 2) if miou is not None else None,
            })

        # Восстанавливаем baseline k
        set_k(model, mtype_ck, baseline_k)

    # ── Консольная сводка ──────────────────────────────────────────────────────
    print("\n\n" + "="*70)
    print("ИТОГОВАЯ ТАБЛИЦА — чувствительность к k")
    print("="*70)
    cur_model = None
    for r in all_results:
        if r["model"] != cur_model:
            cur_model = r["model"]
            print(f"\n  {cur_model} (baseline k={r['baseline_k']})")
            print(f"  {'k':>4}  {'мс':>8}  {'speedup':>8}  {'mIoU%':>8}")
            print(f"  {'-'*36}")
        marker = " *" if r["k"] == r["baseline_k"] else ""
        miou_s = f"{r['miou']:.2f}" if r["miou"] is not None else "—"
        sp_s   = f"{r['speedup']:.2f}x"
        print(f"  {r['k']:>4}  {r['ms']:>8.2f}  {sp_s:>8}  {miou_s:>8}{marker}")

    # ── Excel ──────────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "k sensitivity"

    HDR   = PatternFill("solid", start_color="BDD7EE")
    BASE  = PatternFill("solid", start_color="C6EFCE")   # зелёный = baseline
    thin  = Side(style="thin", color="AAAAAA")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    MODEL_COLORS = ["DEEAF1", "E2EFDA", "FFF2CC", "FCE4D6"]
    model_names = list(dict.fromkeys(r["model"] for r in all_results))
    mc = {m: MODEL_COLORS[i % len(MODEL_COLORS)] for i, m in enumerate(model_names)}

    headers = ["Модель", "k", "Baseline k", "Время (мс)", "Ускорение (x)", "mIoU (%)"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col)
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = brd

    for r_idx, r in enumerate(all_results, 2):
        is_base = r["k"] == r["baseline_k"]
        fill = BASE if is_base else PatternFill("solid", start_color=mc[r["model"]])
        vals = [r["model"], r["k"], r["baseline_k"], r["ms"],
                r["speedup"], r["miou"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(r_idx, col)
            c.value = val if val is not None else "—"
            c.font = Font(bold=is_base, name="Arial", size=10)
            c.fill = fill
            c.border = brd
            c.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEF", [16, 6, 10, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    # ── Лист с данными для графиков ────────────────────────────────────────────
    ws2 = wb.create_sheet("Данные графиков")
    # Скорость: одна колонка на модель
    ws2.cell(1, 1).value = "k"
    col = 2
    col_map = {}
    for m in model_names:
        ws2.cell(1, col).value = m
        col_map[m] = col
        col += 1

    k_all = sorted(set(r["k"] for r in all_results))
    for row_i, k in enumerate(k_all, 2):
        ws2.cell(row_i, 1).value = k
        for m in model_names:
            vals_k = [r["ms"] for r in all_results if r["model"] == m and r["k"] == k]
            ws2.cell(row_i, col_map[m]).value = vals_k[0] if vals_k else None

    chart_speed = LineChart()
    chart_speed.title = "Время инференса vs k"
    chart_speed.y_axis.title = "мс"
    chart_speed.x_axis.title = "k (число соседей)"
    chart_speed.style = 10
    chart_speed.width = 22
    chart_speed.height = 14
    for m in model_names:
        data = Reference(ws2, min_col=col_map[m], min_row=1, max_row=len(k_all)+1)
        chart_speed.add_data(data, titles_from_data=True)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(k_all)+1)
    chart_speed.set_categories(cats)
    ws2.add_chart(chart_speed, "H2")

    # mIoU график (если есть данные)
    if any(r["miou"] is not None for r in all_results):
        col_miou = len(model_names) + 2
        ws2.cell(1, col_miou).value = "k"
        col_m2 = {}
        for m in model_names:
            col_miou += 1
            ws2.cell(1, col_miou).value = f"{m} mIoU"
            col_m2[m] = col_miou

        k_start_row = len(k_all) + 3
        for row_i, k in enumerate(k_all, k_start_row):
            ws2.cell(row_i, len(model_names)+2).value = k
            for m in model_names:
                vals_k = [r["miou"] for r in all_results
                          if r["model"] == m and r["k"] == k and r["miou"] is not None]
                ws2.cell(row_i, col_m2[m]).value = vals_k[0] if vals_k else None

        chart_miou = LineChart()
        chart_miou.title = "mIoU vs k"
        chart_miou.y_axis.title = "mIoU (%)"
        chart_miou.x_axis.title = "k"
        chart_miou.style = 10
        chart_miou.width = 22
        chart_miou.height = 14
        for m in model_names:
            data = Reference(ws2, min_col=col_m2[m], min_row=k_start_row-1,
                             max_row=k_start_row+len(k_all)-1)
            chart_miou.add_data(data, titles_from_data=True)
        cats2 = Reference(ws2, min_col=len(model_names)+2,
                          min_row=k_start_row, max_row=k_start_row+len(k_all)-1)
        chart_miou.set_categories(cats2)
        ws2.add_chart(chart_miou, "H20")

    out = "results/k_sensitivity.xlsx"
    wb.save(out)
    print(f"\nsaved: {out}")


if __name__ == "__main__":
    main()
