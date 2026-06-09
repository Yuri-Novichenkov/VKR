"""
Генерация итоговых таблиц Excel для ВКР.
Объединяет все результаты: точность, инференс GPU/CPU, N/k-sensitivity, INT8.

Запуск:
    python scripts/make_vkr_tables.py
Результат:
    results/vkr_tables.xlsx
"""
import sys, csv, warnings
from pathlib import Path
import numpy as np

sys.path.insert(0, ".")
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ═══════════════════════════════════════════════════════════════════════════════
# ДАННЫЕ
# ═══════════════════════════════════════════════════════════════════════════════

# Порядок моделей для всех таблиц
MODEL_ORDER = [
    "PointNet", "PointNet++", "DGCNN",
    "LDGCNN", "LDGCNN-GATv2", "LDGCNN-LocalWin",
    "LDGCNNFlash", "PointTransformer",
]

# ── 1. Точность моделей (из MLflow + best_params.yaml) ────────────────────────
ACCURACY = {
    "PointNet":        {"miou": 35.62, "acc": 80.88},
    "PointNet++":      {"miou": 44.69, "acc": 85.31},
    "DGCNN":           {"miou": 55.45, "acc": 86.50},
    "LDGCNN":          {"miou": 53.75, "acc": 84.37},
    "LDGCNN-GATv2":    {"miou": 55.05, "acc": 86.77},
    "LDGCNN-LocalWin": {"miou": 55.97, "acc": 87.04},
    "LDGCNNFlash":     {"miou": 60.88, "acc": 87.87},
    "PointTransformer":{"miou": 61.65, "acc": 87.79},
}

# ── 2. Сложность моделей (torchprofile) ───────────────────────────────────────
COMPLEXITY = {
    "PointNet":        {"params": 3_533_204, "gmacs": 4.74},
    "PointNet++":      {"params": 1_404_363, "gmacs": 1.27},
    "DGCNN":           {"params":   258_315, "gmacs": 12.65},
    "LDGCNN":          {"params":   531_979, "gmacs": 32.93},
    "LDGCNN-GATv2":    {"params":   548_491, "gmacs": 34.10},
    "LDGCNN-LocalWin": {"params":   548_363, "gmacs": 34.09},
    "LDGCNNFlash":     {"params":   931_595, "gmacs": 28.63},
    "PointTransformer":{"params": 3_070_347, "gmacs":  3.35},
}

# ── 3. Время обучения (per_class_and_time.txt) ────────────────────────────────
TRAIN_TIME = {
    "PointNet":        {"ep_sec": 142.8, "epochs": 100, "hours": 4.0},
    "PointNet++":      {"ep_sec": 1372.6,"epochs": 100, "hours": 38.1},
    "DGCNN":           {"ep_sec":  37.2, "epochs": 100, "hours": 1.0},
    "LDGCNN":          {"ep_sec":  97.6, "epochs": 100, "hours": 2.7},
    "LDGCNN-GATv2":    {"ep_sec": 110.0, "epochs": 100, "hours": 3.1},
    "LDGCNN-LocalWin": {"ep_sec": 108.0, "epochs": 100, "hours": 3.0},
    "LDGCNNFlash":     {"ep_sec":  38.0, "epochs": 100, "hours": 1.1},
    "PointTransformer":{"ep_sec": 134.1, "epochs": 100, "hours": 3.7},
}

# ── 4. GPU-инференс (benchmark_gpu.csv, A5000, N=4096, batch=1) ───────────────
# Модели в csv: pointnet, pointnet++, dgcnn, ldgcnn(=GATv2), pointtransformer, ldgcnn_flash
GPU_BENCH = {
    "PointNet": {
        "fp32":         {"ms": 2.70,  "vram": 114.4, "thr": 370.4},
        "fp16":         {"ms": 2.87,  "vram":  82.2, "thr": 348.8},
        "compile_fp32": {"ms": 1.51,  "vram":  99.5, "thr": 664.4},
        "compile_fp16": {"ms": 1.53,  "vram":  74.5, "thr": 653.6},
    },
    "PointNet++": {
        "fp32":         {"ms": 229.1, "vram":  87.7, "thr": 4.4},
        "fp16":         {"ms": 228.8, "vram":  84.9, "thr": 4.4},
        "compile_fp32": {"ms": 224.5, "vram":  79.7, "thr": 4.5},
        "compile_fp16": {"ms": 222.4, "vram":  76.9, "thr": 4.5},
    },
    "DGCNN": {
        "fp32":         {"ms":  14.2, "vram": 258.0, "thr":  70.2},
        "fp16":         {"ms":   8.0, "vram": 135.4, "thr": 124.9},
        "compile_fp32": {"ms":   6.3, "vram": 181.2, "thr": 157.5},
        "compile_fp16": {"ms":   3.7, "vram":  96.7, "thr": 269.1},
    },
    "LDGCNN-GATv2": {
        "fp32":         {"ms":  37.9, "vram": 508.2, "thr":  26.4},
        "fp16":         {"ms":  21.5, "vram": 263.8, "thr":  46.5},
        "compile_fp32": {"ms":  28.7, "vram": 510.5, "thr":  34.8},
        "compile_fp16": {"ms":  17.1, "vram": 262.6, "thr":  58.5},
    },
    "LDGCNNFlash": {
        "fp32":         {"ms":  14.9, "vram": 379.6, "thr":  67.1},
        "fp16":         {"ms":   8.2, "vram": 199.6, "thr": 122.5},
        "compile_fp32": {"ms":  12.7, "vram": 282.1, "thr":  78.9},
        "compile_fp16": {"ms":   5.9, "vram": 155.8, "thr": 168.2},
    },
    "PointTransformer": {
        "fp32":         {"ms": 169.0, "vram": 187.2, "thr":   5.9},
        "fp16":         {"ms": 174.2, "vram": 180.3, "thr":   5.7},
        "compile_fp32": {"ms":   8.7, "vram": 143.8, "thr": 114.8},
        "compile_fp16": {"ms":   9.2, "vram": 147.8, "thr": 109.0},
    },
}
GPU_BEST = {  # лучший режим
    "PointNet":         ("compile_fp32", 1.51, 1.79),
    "PointNet++":       ("compile_fp16", 222.4, 1.03),
    "DGCNN":            ("compile_fp16", 3.72, 3.83),
    "LDGCNN-GATv2":     ("compile_fp16", 17.1, 2.22),
    "LDGCNNFlash":      ("compile_fp16", 5.95, 2.50),
    "PointTransformer": ("compile_fp32", 8.71, 19.41),
}

# ── 5. CPU-инференс (benchmark_cpu.csv, N=4096) ───────────────────────────────
CPU_BENCH_N4096 = {
    "PointNet":         {"fp32": 74.9,   "onnx":  56.2},
    "PointNet++":       {"fp32": 300.3,  "onnx":   None},   # ONNX не работает
    "DGCNN":            {"fp32": 509.1,  "onnx":  321.1},
    "LDGCNN-GATv2":     {"fp32": 1374.6, "onnx":  771.0},
    "LDGCNNFlash":      {"fp32": 484.3,  "onnx":  442.9},
    "PointTransformer": {"fp32": 553.5,  "onnx":  616.8},   # ONNX медленнее!
}

# ── 6. N-sensitivity CPU (n_sensitivity_cpu.xlsx) ─────────────────────────────
N_CPU = {
    "PointNet":         {512:{"ms":10.6,"miou":32.54}, 1024:{"ms":16.9,"miou":36.57},
                         2048:{"ms":32.7,"miou":38.21}, 3072:{"ms":52.9,"miou":38.34},
                         4096:{"ms":63.9,"miou":37.53}},
    "PointNet++":       {512:{"ms":128.0,"miou":35.54}, 1024:{"ms":220.9,"miou":39.75},
                         2048:{"ms":215.5,"miou":43.44}, 3072:{"ms":409.9,"miou":43.50},
                         4096:{"ms":306.0,"miou":43.44}},
    "DGCNN":            {512:{"ms":34.5,"miou":39.21}, 1024:{"ms":79.9,"miou":45.10},
                         2048:{"ms":168.3,"miou":47.19}, 3072:{"ms":308.3,"miou":47.05},
                         4096:{"ms":465.5,"miou":46.57}},
    "LDGCNN":           {512:{"ms":92.9,"miou":37.44}, 1024:{"ms":213.2,"miou":44.92},
                         2048:{"ms":512.5,"miou":48.93}, 3072:{"ms":815.8,"miou":49.83},
                         4096:{"ms":1224.6,"miou":49.75}},
    "LDGCNN-GATv2":     {512:{"ms":140.1,"miou":38.56}, 1024:{"ms":241.7,"miou":46.08},
                         2048:{"ms":618.6,"miou":49.59}, 3072:{"ms":910.3,"miou":50.17},
                         4096:{"ms":1298.9,"miou":49.88}},
    "LDGCNN-LocalWin":  {512:{"ms":84.4,"miou":37.57}, 1024:{"ms":222.2,"miou":44.99},
                         2048:{"ms":583.4,"miou":49.52}, 3072:{"ms":856.7,"miou":50.05},
                         4096:{"ms":1411.6,"miou":49.90}},
    "LDGCNNFlash":      {512:{"ms":35.3,"miou":45.59}, 1024:{"ms":63.3,"miou":52.50},
                         2048:{"ms":188.5,"miou":55.20}, 3072:{"ms":328.3,"miou":56.17},
                         4096:{"ms":461.5,"miou":56.27}},
    "PointTransformer": {512:{"ms":196.4,"miou":11.18}, 1024:{"ms":225.1,"miou":18.56},
                         2048:{"ms":314.1,"miou":40.55}, 3072:{"ms":412.3,"miou":47.81},
                         4096:{"ms":521.5,"miou":49.28}},
}

# ── 7. N-sensitivity GPU (n_sensitivity_cuda.xlsx) ────────────────────────────
N_GPU = {
    "PointNet":         {512:{"ms":2.8,"miou":32.59}, 1024:{"ms":2.6,"miou":36.57},
                         2048:{"ms":2.6,"miou":38.21}, 3072:{"ms":2.7,"miou":38.33},
                         4096:{"ms":2.7,"miou":37.53}},
    "PointNet++":       {512:{"ms":218.7,"miou":35.62}, 1024:{"ms":224.9,"miou":39.74},
                         2048:{"ms":225.0,"miou":43.61}, 3072:{"ms":224.0,"miou":43.73},
                         4096:{"ms":219.9,"miou":43.55}},
    "DGCNN":            {512:{"ms":2.7,"miou":39.26}, 1024:{"ms":3.1,"miou":45.22},
                         2048:{"ms":4.8,"miou":47.19}, 3072:{"ms":9.1,"miou":47.07},
                         4096:{"ms":14.3,"miou":46.57}},
    "LDGCNN":           {512:{"ms":5.5,"miou":37.26}, 1024:{"ms":6.7,"miou":45.14},
                         2048:{"ms":14.2,"miou":48.90}, 3072:{"ms":24.2,"miou":49.80},
                         4096:{"ms":35.6,"miou":49.75}},
    "LDGCNN-GATv2":     {512:{"ms":6.5,"miou":38.50}, 1024:{"ms":7.9,"miou":45.97},
                         2048:{"ms":15.1,"miou":49.60}, 3072:{"ms":25.8,"miou":50.19},
                         4096:{"ms":38.1,"miou":49.81}},
    "LDGCNN-LocalWin":  {512:{"ms":6.5,"miou":37.57}, 1024:{"ms":7.7,"miou":45.06},
                         2048:{"ms":15.0,"miou":49.56}, 3072:{"ms":25.6,"miou":50.07},
                         4096:{"ms":37.9,"miou":49.91}},
    "LDGCNNFlash":      {512:{"ms":3.0,"miou":45.68}, 1024:{"ms":3.5,"miou":52.54},
                         2048:{"ms":5.5,"miou":55.19}, 3072:{"ms":9.9,"miou":56.29},
                         4096:{"ms":15.2,"miou":56.27}},
    "PointTransformer": {512:{"ms":166.2,"miou":11.19}, 1024:{"ms":163.8,"miou":18.73},
                         2048:{"ms":163.1,"miou":40.46}, 3072:{"ms":162.5,"miou":47.87},
                         4096:{"ms":161.2,"miou":49.03}},
}

# ── 8. k-sensitivity GPU ──────────────────────────────────────────────────────
K_SENS = {
    "DGCNN": {
        5:{"ms":8.7,"miou":31.27}, 8:{"ms":9.4,"miou":37.98},
        10:{"ms":9.8,"miou":41.13}, 16:{"ms":12.5,"miou":45.47},
        20:{"ms":14.4,"miou":46.57}, 24:{"ms":15.9,"miou":47.14},
        32:{"ms":18.8,"miou":47.03},
    },
    "LDGCNN": {
        5:{"ms":18.5,"miou":36.20}, 8:{"ms":21.9,"miou":43.03},
        10:{"ms":24.1,"miou":45.57}, 16:{"ms":30.8,"miou":49.33},
        20:{"ms":35.7,"miou":49.75}, 24:{"ms":40.3,"miou":49.40},
        32:{"ms":48.6,"miou":47.61},
    },
    "LDGCNNFlash": {
        5:{"ms":12.3,"miou":41.63}, 8:{"ms":12.9,"miou":48.03},
        10:{"ms":13.3,"miou":50.66}, 16:{"ms":14.4,"miou":55.34},
        20:{"ms":15.2,"miou":56.27}, 24:{"ms":15.9,"miou":56.49},
        32:{"ms":17.4,"miou":55.71},
    },
    "PointTransformer": {
        4:{"ms":164.7,"miou":36.92}, 8:{"ms":166.9,"miou":47.39},
        12:{"ms":167.3,"miou":49.72}, 16:{"ms":169.7,"miou":49.07},
        20:{"ms":167.1,"miou":47.32}, 24:{"ms":167.4,"miou":45.02},
        32:{"ms":170.0,"miou":39.24},
    },
}
K_BASELINE = {"DGCNN": 20, "LDGCNN": 20, "LDGCNNFlash": 20, "PointTransformer": 16}

# ── 9. INT8 CPU (optimize_benchmark.py, N=1024) ───────────────────────────────
INT8 = {
    "PointNet":         {"fp32_cpu": 19.5, "int8_cpu": 19.4, "speedup": 1.00},
    "PointNet++":       {"fp32_cpu": 177.2,"int8_cpu": 176.8,"speedup": 1.00},
    "DGCNN":            {"fp32_cpu": 90.6, "int8_cpu": 94.8, "speedup": 0.96},
    "LDGCNN":           {"fp32_cpu": 247.3,"int8_cpu": 245.6,"speedup": 1.01},
    "LDGCNN-GATv2":     {"fp32_cpu": 259.3,"int8_cpu": 268.1,"speedup": 0.97},
    "LDGCNN-LocalWin":  {"fp32_cpu": 264.4,"int8_cpu": 254.0,"speedup": 1.04},
    "LDGCNNFlash":      {"fp32_cpu": 83.2, "int8_cpu": 80.7, "speedup": 1.03},
    "PointTransformer": {"fp32_cpu": 271.8,"int8_cpu": 272.2,"speedup": 1.00},
}

N_VALUES = [512, 1024, 2048, 3072, 4096]

# ═══════════════════════════════════════════════════════════════════════════════
# СТИЛИ
# ═══════════════════════════════════════════════════════════════════════════════

# Цвета моделей
MODEL_COLORS = {
    "PointNet":         "DEEAF1",
    "PointNet++":       "E2EFDA",
    "DGCNN":            "FFF2CC",
    "LDGCNN":           "FCE4D6",
    "LDGCNN-GATv2":     "F4CCFF",
    "LDGCNN-LocalWin":  "FFD9E8",
    "LDGCNNFlash":      "D5F5E3",
    "PointTransformer": "D6EAF8",
}

C_HDR    = "1F3864"  # тёмно-синий заголовок
C_HDR2   = "2E75B6"  # средний синий подзаголовок
C_GOLD   = "FFD700"  # лучший результат
C_GREEN  = "C6EFCE"  # хорошо
C_ORANGE = "FFDAB9"  # плохо
C_WHITE  = "FFFFFF"

def hdr_style(ws, row, col, value, level=1, colspan=1, wrap=True):
    c = ws.cell(row, col, value)
    bg = C_HDR if level == 1 else C_HDR2
    c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(bold=True, color="FFFFFF", name="Calibri",
                  size=11 if level == 1 else 10)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=wrap)
    _border(c)
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    return c

def data_cell(ws, row, col, value, model=None, bold=False, fmt=None,
              align="center", color=None):
    c = ws.cell(row, col, value)
    bg = color or (MODEL_COLORS.get(model, C_WHITE) if model else C_WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(bold=bold, name="Calibri", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    _border(c)
    if fmt and isinstance(value, (int, float)):
        c.number_format = fmt
    return c

def _border(c):
    thin = Side(style="thin", color="AAAAAA")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def row_h(ws, row, h=18):
    ws.row_dimensions[row].height = h

def highlight_best_col(ws, data_rows, col, model_col=1, higher_better=True):
    """Подсвечивает лучшее значение в столбце золотым, худшее — оранжевым."""
    vals = [(ws.cell(r, col).value, r) for r in data_rows
            if isinstance(ws.cell(r, col).value, (int, float))]
    if not vals:
        return
    vals.sort(key=lambda x: x[0], reverse=higher_better)
    best_r = vals[0][1]
    worst_r = vals[-1][1]
    ws.cell(best_r, col).fill = PatternFill("solid", start_color=C_GOLD)
    ws.cell(best_r, col).font = Font(bold=True, name="Calibri", size=10)
    if worst_r != best_r:
        ws.cell(worst_r, col).fill = PatternFill("solid", start_color=C_ORANGE)

# ═══════════════════════════════════════════════════════════════════════════════
# ЛИСТЫ
# ═══════════════════════════════════════════════════════════════════════════════

wb = Workbook()
wb.remove(wb.active)  # удаляем пустой лист

# ─────────────────────────────────────────────────────────────────────────────
# Лист 1: Сравнение моделей (главная таблица)
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("1. Сравнение моделей")
ws1.freeze_panes = "B3"
ws1.sheet_view.showGridLines = False

TITLE = "Сравнение архитектур нейронных сетей для семантической сегментации (датасет Hessigheim Mar16)"
ws1.merge_cells("A1:N1")
t = ws1.cell(1, 1, TITLE)
t.fill = PatternFill("solid", start_color=C_HDR)
t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws1, 1, 24)

# Заголовки групп
hdr_style(ws1, 2, 1,  "Модель",           level=1)
hdr_style(ws1, 2, 2,  "Архитектура",      level=1, colspan=2)
hdr_style(ws1, 2, 4,  "Точность (Mar16)", level=1, colspan=2)
hdr_style(ws1, 2, 6,  "Обучение (GPU A5000)", level=1, colspan=3)
hdr_style(ws1, 2, 9,  "Инференс GPU (A5000, N=4096, batch=1)", level=1, colspan=4)
hdr_style(ws1, 2, 13, "Инференс CPU (i7-10750H)", level=1, colspan=2)
row_h(ws1, 2, 20)

# Подзаголовки
subs = [
    "Параметры", "GMACs",
    "mIoU (%)", "Точность (%)",
    "Время эпохи (с)", "Эпох", "Всего (ч)",
    "FP32 (мс)", "FP16 (мс)", "Лучший (мс)", "Режим",
    "FP32, N=4096 (мс)", "ONNX, N=4096 (мс)",
]
for i, s in enumerate(subs, 2):
    hdr_style(ws1, 3, i, s, level=2, wrap=True)
    row_h(ws1, 3, 30)
hdr_style(ws1, 3, 1, "Модель", level=2)

data_rows_s1 = []
for ri, model in enumerate(MODEL_ORDER, 4):
    row_h(ws1, ri, 18)
    data_rows_s1.append(ri)
    data_cell(ws1, ri, 1,  model, model=model, bold=True, align="left")
    c = COMPLEXITY[model]
    data_cell(ws1, ri, 2,  c["params"], model=model, fmt="#,##0")
    data_cell(ws1, ri, 3,  c["gmacs"],  model=model, fmt="0.00")
    a = ACCURACY[model]
    data_cell(ws1, ri, 4,  a["miou"],   model=model, fmt="0.00")
    data_cell(ws1, ri, 5,  a["acc"],    model=model, fmt="0.00")
    t = TRAIN_TIME[model]
    data_cell(ws1, ri, 6,  t["ep_sec"], model=model, fmt="0.0")
    data_cell(ws1, ri, 7,  t["epochs"], model=model)
    data_cell(ws1, ri, 8,  t["hours"],  model=model, fmt="0.0")
    # GPU bench
    gb = GPU_BENCH.get(model, {})
    fp32_ms = gb.get("fp32", {}).get("ms")
    fp16_ms = gb.get("fp16", {}).get("ms")
    best_info = GPU_BEST.get(model)
    data_cell(ws1, ri, 9,  fp32_ms,                      model=model, fmt="0.0")
    data_cell(ws1, ri, 10, fp16_ms,                      model=model, fmt="0.0")
    data_cell(ws1, ri, 11, best_info[1] if best_info else None, model=model, fmt="0.0")
    data_cell(ws1, ri, 12, best_info[0] if best_info else "—",  model=model)
    # CPU bench
    cb = CPU_BENCH_N4096.get(model, {})
    data_cell(ws1, ri, 13, cb.get("fp32"),  model=model, fmt="0.0")
    onnx_v = cb.get("onnx")
    data_cell(ws1, ri, 14, onnx_v if onnx_v else "—", model=model, fmt="0.0")

highlight_best_col(ws1, data_rows_s1, 4)   # mIoU
highlight_best_col(ws1, data_rows_s1, 5)   # точность
highlight_best_col(ws1, data_rows_s1, 11, higher_better=False)  # лучший GPU
highlight_best_col(ws1, data_rows_s1, 13, higher_better=False)  # CPU fp32

widths1 = [18, 13, 8, 10, 11, 14, 7, 9, 11, 11, 12, 14, 15, 15]
for i, w in enumerate(widths1, 1):
    col_width(ws1, i, w)

# Примечание
note_row = len(MODEL_ORDER) + 5
ws1.cell(note_row, 1, "* GPU-бенчмарк: NVIDIA RTX A5000, batch=1, N=4096. "
         "LDGCNN в GPU-бенчмарке — вариант с GATv2. "
         "PointNet++ ONNX: экспорт завершился с ошибкой (динамические формы в index_put_). "
         "Золото = лучший результат, оранжевый = худший.").font = Font(italic=True, size=9, name="Calibri")
ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=14)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 2: GPU-оптимизация (все режимы)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("2. GPU-оптимизация")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:M1")
t2 = ws2.cell(1, 1, "Оптимизация инференса на GPU (RTX A5000, N=4096, batch=1)")
t2.fill = PatternFill("solid", start_color=C_HDR)
t2.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t2.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws2, 1, 24)

hdr_style(ws2, 2, 1, "Модель",                  level=1)
hdr_style(ws2, 2, 2, "FP32",                    level=1, colspan=3)
hdr_style(ws2, 2, 5, "FP16",                    level=1, colspan=3)
hdr_style(ws2, 2, 8, "compile + FP32",           level=1, colspan=3)
hdr_style(ws2, 2, 11,"compile + FP16",           level=1, colspan=3)
row_h(ws2, 2, 20)

mode_subs = ["Задержка (мс)", "VRAM (МБ)", "Прокиз./с"]
for gi, label in enumerate(["FP32","FP16","compile FP32","compile FP16"]):
    for j, sub in enumerate(mode_subs):
        hdr_style(ws2, 3, 2 + gi*3 + j, sub, level=2, wrap=True)
hdr_style(ws2, 3, 1, "Модель", level=2)
row_h(ws2, 3, 30)

mode_keys = ["fp32", "fp16", "compile_fp32", "compile_fp16"]
data_rows_s2 = []
for ri, model in enumerate(MODEL_ORDER, 4):
    row_h(ws2, ri, 18)
    data_rows_s2.append(ri)
    data_cell(ws2, ri, 1, model, model=model, bold=True, align="left")
    gb = GPU_BENCH.get(model, {})
    for gi, mk in enumerate(mode_keys):
        d = gb.get(mk, {})
        ms   = d.get("ms")
        vram = d.get("vram")
        thr  = d.get("thr")
        data_cell(ws2, ri, 2 + gi*3,     ms,   model=model, fmt="0.0")
        data_cell(ws2, ri, 2 + gi*3 + 1, vram, model=model, fmt="0")
        data_cell(ws2, ri, 2 + gi*3 + 2, thr,  model=model, fmt="0.0")

# Подсветка лучших задержек по каждому режиму
for gi in range(4):
    highlight_best_col(ws2, data_rows_s2, 2 + gi*3, higher_better=False)

# Доп. строка: ускорение compile_fp16 vs fp32
acc_row = len(MODEL_ORDER) + 5
data_cell(ws2, acc_row, 1, "Ускорение compile_fp16 vs FP32:", bold=True)
for ri, model in enumerate(MODEL_ORDER, acc_row):
    gb = GPU_BENCH.get(model, {})
    fp32 = gb.get("fp32", {}).get("ms")
    best = gb.get("compile_fp16", {}).get("ms")
    if fp32 and best:
        x = round(fp32/best, 2)
        c = data_cell(ws2, acc_row, ri - acc_row + 2, f"{x}x")
        if x >= 3:
            c.fill = PatternFill("solid", start_color=C_GOLD)
            c.font = Font(bold=True, name="Calibri", size=10)

widths2 = [18] + [11, 10, 11]*4
for i, w in enumerate(widths2, 1):
    col_width(ws2, i, w)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 3: N-sensitivity
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("3. N-sensitivity")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:R1")
t3 = ws3.cell(1, 1, "Чувствительность к числу точек N: скорость инференса и mIoU")
t3.fill = PatternFill("solid", start_color=C_HDR)
t3.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t3.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws3, 1, 24)

# Две подтаблицы: CPU слева, GPU справа
hdr_style(ws3, 2, 1, "CPU (i7-10750H)", level=1, colspan=9)
hdr_style(ws3, 2, 11,"GPU (RTX A5000)",  level=1, colspan=9)
row_h(ws3, 2, 20)

hdr_style(ws3, 3, 1,  "Модель",   level=2)
hdr_style(ws3, 3, 11, "Модель",   level=2)
for i, n in enumerate(N_VALUES):
    hdr_style(ws3, 3, 2  + i*2 - 1, f"N={n}\n(мс)",   level=2, wrap=True)
    hdr_style(ws3, 3, 2  + i*2,     f"N={n}\nmIoU%",  level=2, wrap=True)
    hdr_style(ws3, 3, 12 + i*2 - 1, f"N={n}\n(мс)",   level=2, wrap=True)
    hdr_style(ws3, 3, 12 + i*2,     f"N={n}\nmIoU%",  level=2, wrap=True)
row_h(ws3, 3, 32)

data_rows_s3 = []
for ri, model in enumerate(MODEL_ORDER, 4):
    row_h(ws3, ri, 18)
    data_rows_s3.append(ri)
    data_cell(ws3, ri, 1,  model, model=model, bold=True, align="left")
    data_cell(ws3, ri, 11, model, model=model, bold=True, align="left")
    for i, n in enumerate(N_VALUES):
        cpu_d = N_CPU.get(model, {}).get(n, {})
        gpu_d = N_GPU.get(model, {}).get(n, {})
        data_cell(ws3, ri, 2  + i*2 - 1, cpu_d.get("ms"),   model=model, fmt="0.0")
        data_cell(ws3, ri, 2  + i*2,     cpu_d.get("miou"),  model=model, fmt="0.00")
        data_cell(ws3, ri, 12 + i*2 - 1, gpu_d.get("ms"),   model=model, fmt="0.0")
        data_cell(ws3, ri, 12 + i*2,     gpu_d.get("miou"),  model=model, fmt="0.00")

# Подсветка лучших mIoU по каждой N-колонке
for i in range(len(N_VALUES)):
    highlight_best_col(ws3, data_rows_s3, 2  + i*2)      # CPU mIoU
    highlight_best_col(ws3, data_rows_s3, 12 + i*2)      # GPU mIoU

widths3 = [18] + [9,8]*5 + [2] + [18] + [9,8]*5
for i, w in enumerate(widths3, 1):
    col_width(ws3, i, w)
col_width(ws3, 10, 3)

# Примечания
note3 = len(MODEL_ORDER) + 6
ws3.cell(note3, 1, "* CPU: N=4096 не запускался для тяжёлых моделей (зависание O(N^2)). "
                    "Красный: PointTransformer при N<=1024 — коллапс mIoU из-за иерархического FPS.").font = Font(italic=True, size=9, name="Calibri")
ws3.merge_cells(start_row=note3, start_column=1, end_row=note3, end_column=19)

# Выделяем PointTransformer плохие результаты при малом N
for ri, model in enumerate(MODEL_ORDER, 4):
    if model == "PointTransformer":
        for i in range(2):  # N=512, N=1024
            for dc in [2+i*2, 12+i*2]:  # cpu miou, gpu miou
                c = ws3.cell(ri, dc)
                if isinstance(c.value, (int, float)) and c.value < 25:
                    c.fill = PatternFill("solid", start_color="FF9999")
                    c.font = Font(bold=True, name="Calibri", size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 4: k-sensitivity
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("4. k-sensitivity")
ws4.sheet_view.showGridLines = False

k_models = ["DGCNN", "LDGCNN", "LDGCNNFlash", "PointTransformer"]
all_k_vals = sorted({k for m in K_SENS.values() for k in m.keys()})
# col layout: 1=Model, 2=Baseline k, 3..=(ms,mIoU) pairs per k
_last_k_col = 2 + len(all_k_vals) * 2
ws4.merge_cells(f"A1:{get_column_letter(_last_k_col)}1")
t4 = ws4.cell(1, 1, "Чувствительность к числу соседей k: скорость и mIoU (GPU RTX A5000, N=4096)")
t4.fill = PatternFill("solid", start_color=C_HDR)
t4.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t4.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws4, 1, 24)

hdr_style(ws4, 2, 1, "Модель",  level=2)
hdr_style(ws4, 2, 2, "Base. k", level=2)
for i, k in enumerate(all_k_vals):
    hdr_style(ws4, 2, 3 + i*2,   f"k={k}\n(мс)",  level=2, wrap=True)
    hdr_style(ws4, 2, 3 + i*2+1, f"k={k}\nmIoU%", level=2, wrap=True)
row_h(ws4, 2, 32)

data_rows_s4 = []
for ri, model in enumerate(k_models, 3):
    row_h(ws4, ri, 18)
    data_rows_s4.append(ri)
    data_cell(ws4, ri, 1, model, model=model, bold=True, align="left")
    data_cell(ws4, ri, 2, K_BASELINE[model], model=model)
    for i, k in enumerate(all_k_vals):
        d = K_SENS.get(model, {}).get(k, {})
        ms_c  = data_cell(ws4, ri, 3 + i*2,   d.get("ms"),   model=model, fmt="0.0")
        mio_c = data_cell(ws4, ri, 3 + i*2+1, d.get("miou"), model=model, fmt="0.00")
        # Отметить baseline
        if k == K_BASELINE.get(model):
            ms_c.font  = Font(bold=True, name="Calibri", size=10, color="1F3864")
            mio_c.font = Font(bold=True, name="Calibri", size=10, color="1F3864")

# Подсветка лучших mIoU по строке (по каждой модели)
for ri, model in zip(data_rows_s4, k_models):
    mious = [(ws4.cell(ri, 3+i*2+1).value, 3+i*2+1) for i in range(len(all_k_vals))
             if isinstance(ws4.cell(ri, 3+i*2+1).value, (int, float))]
    if mious:
        best_col = max(mious)[1]
        ws4.cell(ri, best_col).fill = PatternFill("solid", start_color=C_GOLD)
        ws4.cell(ri, best_col).font = Font(bold=True, name="Calibri", size=10)

widths4 = [18, 8] + [9, 8]*len(all_k_vals)
for i, w in enumerate(widths4, 1):
    col_width(ws4, i, w)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 5: INT8 квантизация
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("5. INT8 квантизация")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:F1")
t5 = ws5.cell(1, 1, "Dynamic INT8 квантизация на CPU (i7-10750H, N=1024)")
t5.fill = PatternFill("solid", start_color=C_HDR)
t5.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t5.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws5, 1, 24)

for ci, hdr in enumerate(["Модель","FP32 CPU (мс)","INT8 CPU (мс)","Ускорение (x)","Вывод"], 1):
    hdr_style(ws5, 2, ci, hdr, level=2)
row_h(ws5, 2, 22)

for ri, model in enumerate(MODEL_ORDER, 3):
    row_h(ws5, ri, 18)
    d = INT8[model]
    data_cell(ws5, ri, 1, model, model=model, bold=True, align="left")
    data_cell(ws5, ri, 2, d["fp32_cpu"], model=model, fmt="0.0")
    data_cell(ws5, ri, 3, d["int8_cpu"], model=model, fmt="0.0")
    sp = d["speedup"]
    c_sp = data_cell(ws5, ri, 4, sp, model=model, fmt="0.00")
    if sp < 1.0:
        c_sp.fill = PatternFill("solid", start_color=C_ORANGE)
        verdict = "Замедление"
    elif sp < 1.05:
        verdict = "Без эффекта"
    else:
        verdict = "Минимальное ускорение"
    data_cell(ws5, ri, 5, verdict, model=model)

ws5.cell(len(MODEL_ORDER)+5, 1,
    "Вывод: Dynamic INT8 quantization не даёт ускорения для граф-моделей. "
    "Узкое место — pairwise_distance (O(N^2) matmul) — не является квантизируемым слоем. "
    "Conv2d в EdgeConv занимают лишь ~10-15% времени вычислений."
).font = Font(italic=True, size=9, name="Calibri")
ws5.merge_cells(start_row=len(MODEL_ORDER)+5, start_column=1,
                end_row=len(MODEL_ORDER)+5, end_column=5)

for i, w in enumerate([18, 14, 14, 14, 22], 1):
    col_width(ws5, i, w)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 6: Оптимальный N — рекомендации
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("6. Оптимальный N")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:I1")
t6 = ws6.cell(1, 1, "Рекомендации по выбору N: лучший компромисс скорость/точность")
t6.fill = PatternFill("solid", start_color=C_HDR)
t6.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t6.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws6, 1, 24)

hdrs6 = ["Модель","Opt. N (CPU)","мс CPU","mIoU% @opt.N","mIoU% @4096",
          "Потеря mIoU","Ускорение (x)","Opt. N (GPU)","мс GPU"]
for ci, h in enumerate(hdrs6, 1):
    hdr_style(ws6, 2, ci, h, level=2, wrap=True)
row_h(ws6, 2, 30)

# Лучший N для CPU = минимальный N где mIoU >= baseline_mIoU - 3%
def best_n_cpu(model):
    base_miou = N_CPU[model][4096]["miou"]
    threshold = base_miou - 3.0
    for n in N_VALUES:
        d = N_CPU[model].get(n, {})
        if d.get("miou", 0) >= threshold:
            return n, d["ms"], d["miou"]
    return 4096, N_CPU[model][4096]["ms"], base_miou

def best_n_gpu(model):
    # GPU — ищем N где mIoU >= baseline - 2%
    base_miou = N_GPU[model][4096]["miou"]
    threshold = base_miou - 2.0
    for n in N_VALUES:
        d = N_GPU[model].get(n, {})
        if d.get("miou", 0) >= threshold:
            return n, d["ms"]
    return 4096, N_GPU[model][4096]["ms"]

for ri, model in enumerate(MODEL_ORDER, 3):
    row_h(ws6, ri, 18)
    opt_n_cpu, ms_cpu, miou_opt = best_n_cpu(model)
    opt_n_gpu, ms_gpu           = best_n_gpu(model)
    miou_base = N_CPU[model][4096]["miou"]
    loss = round(miou_base - miou_opt, 2)
    spd  = round(N_CPU[model][4096]["ms"] / ms_cpu, 2)

    data_cell(ws6, ri, 1, model,    model=model, bold=True, align="left")
    data_cell(ws6, ri, 2, opt_n_cpu,model=model)
    data_cell(ws6, ri, 3, ms_cpu,   model=model, fmt="0.0")
    data_cell(ws6, ri, 4, miou_opt, model=model, fmt="0.00")
    data_cell(ws6, ri, 5, miou_base,model=model, fmt="0.00")
    c_loss = data_cell(ws6, ri, 6, loss if loss > 0 else 0.0, model=model, fmt="0.00")
    if loss < 0:
        c_loss.fill = PatternFill("solid", start_color=C_GREEN)
    elif loss > 5:
        c_loss.fill = PatternFill("solid", start_color=C_ORANGE)
    c_spd = data_cell(ws6, ri, 7, spd, model=model, fmt="0.00")
    if spd >= 5:
        c_spd.fill = PatternFill("solid", start_color=C_GOLD)
        c_spd.font = Font(bold=True, name="Calibri", size=10)
    data_cell(ws6, ri, 8, opt_n_gpu, model=model)
    data_cell(ws6, ri, 9, ms_gpu,    model=model, fmt="0.0")

for i, w in enumerate([18,11,11,12,12,12,13,11,10], 1):
    col_width(ws6, i, w)

ws6.cell(len(MODEL_ORDER)+5, 1,
    "* Оптимальный N: минимальный при котором mIoU >= (mIoU@4096 - 3%) для CPU, -2% для GPU. "
    "PointTransformer требует N>=2048 для работоспособности из-за иерархического FPS."
).font = Font(italic=True, size=9, name="Calibri")
ws6.merge_cells(start_row=len(MODEL_ORDER)+5, start_column=1,
                end_row=len(MODEL_ORDER)+5, end_column=9)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 7: Сводка методов оптимизации
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("7. Методы оптимизации")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:G1")
t7 = ws7.cell(1, 1, "Сводная таблица методов оптимизации инференса")
t7.fill = PatternFill("solid", start_color=C_HDR)
t7.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
t7.alignment = Alignment(horizontal="center", vertical="center")
row_h(ws7, 1, 24)

hdrs7 = ["Метод","Устройство","Применимость","Макс. ускорение","Потеря точности","Реализация","Вывод"]
for ci, h in enumerate(hdrs7, 1):
    hdr_style(ws7, 2, ci, h, level=2, wrap=True)
row_h(ws7, 2, 24)

methods = [
    ("FP16 (half precision)",
     "GPU", "Все модели кроме PointNet++ (NaN)",
     "3.8x (DGCNN)", "Нет",
     "model.half() или AMP",
     "Рекомендуется для GPU"),
    ("torch.compile (inductor)",
     "GPU", "Все модели",
     "19.4x (PointTransformer!)", "Нет",
     "torch.compile(model)",
     "Рекомендуется; PointTransformer — исключительный выигрыш"),
    ("compile + FP16",
     "GPU", "Все кроме PointNet++ и PointTransformer",
     "3.8x (DGCNN)", "Нет",
     "model.half() + torch.compile",
     "Лучший режим для DGCNN, LDGCNN, LDGCNNFlash"),
    ("ONNX Runtime (CPU)",
     "CPU", "Все кроме PointNet++ (ошибка экспорта)",
     "1.7x (PointNet)", "Нет",
     "torch.onnx.export + onnxruntime",
     "Умеренный эффект; PointTransformer — медленнее FP32"),
    ("Уменьшение N (num_points)",
     "CPU + GPU", "Все модели (PT: N>=2048)",
     "15.4x CPU (DGCNN N=512)", "3-8% mIoU при N=1024",
     "Параметр --num_points в dataset",
     "Главный инструмент ускорения на CPU"),
    ("Уменьшение k (k-соседей)",
     "GPU", "DGCNN, LDGCNN, LDGCNNFlash",
     "1.9x (DGCNN k=5 vs k=20)", "5-15% mIoU при k=5",
     "Параметр k при создании модели",
     "Умеренный эффект; PT нечувствителен к k"),
    ("Dynamic INT8 quantization",
     "CPU", "Все модели",
     "1.04x (незначимо)", "Нет",
     "torch.quantization.quantize_dynamic",
     "Не применимо: узкое место — matmul kNN, не Conv2d"),
    ("fast_kNN (torch_cluster)",
     "GPU", "Граф-модели (DGCNN, LDGCNN, Flash)",
     "Медленнее (~14x) при N=4096", "Нет",
     "fast_knn=True при построении модели",
     "Не рекомендуется: cuBLAS GEMM эффективнее для плотных батчей"),
]

MCOLORS = ["E8F5E9","E3F2FD","FFF8E1","F3E5F5","FBE9E7","E0F7FA","FCE4EC","F1F8E9"]
for ri, (meth, dev, appl, spd, loss, impl, concl) in enumerate(methods, 3):
    row_h(ws7, ri, 28)
    bg = MCOLORS[(ri-3) % len(MCOLORS)]
    fill = PatternFill("solid", start_color=bg)
    for ci, val in enumerate([meth, dev, appl, spd, loss, impl, concl], 1):
        c = ws7.cell(ri, ci, val)
        c.fill = fill
        c.font = Font(name="Calibri", size=10,
                      bold=(ci==1), color="CC0000" if "Не" in concl and ci==7 else "000000")
        c.alignment = Alignment(horizontal="left" if ci > 2 else "center",
                                vertical="center", wrap_text=True)
        _border(c)

for i, w in enumerate([24,12,28,22,18,28,34], 1):
    col_width(ws7, i, w)


# ═══════════════════════════════════════════════════════════════════════════════
# СОХРАНЕНИЕ
# ═══════════════════════════════════════════════════════════════════════════════
Path("results").mkdir(exist_ok=True)
out = "results/vkr_tables.xlsx"
wb.save(out)
print(f"Сохранено: {out}")
print(f"Листы: {', '.join(s.title for s in wb.worksheets)}")
