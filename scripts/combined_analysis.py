"""
Объединённый анализ: MACs по слоям (torchinfo) + время по слоям (CUDA hooks).
Сохраняет Excel с таблицей и диаграммами.
"""
import sys, torch, numpy as np
from pathlib import Path

sys.path.insert(0, ".")
from src.models import build_model

CONFIGS = [
    ("PointNet",         "pointnet",         "checkpoints/loss_sweep/pointnet/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointNet++",       "pointnet++",       "checkpoints/loss_sweep/pointnet++/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("DGCNN",            "dgcnn",            "checkpoints/loss_sweep/dgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN",           "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN GAT",       "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_gatv2_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN WINDOW",    "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_local_window_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNNFlash",      "ldgcnn_flash",     "checkpoints/loss_sweep/ldgcnn_flash/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointTransformer", "pointtransformer", "checkpoints/pointtransformer/segmentation/cb_effective_b0p999/mar16/best_model.pth"),
]
DEFAULTS = dict(k=20, k_small=20, k_large=40, attention_type="none",
                attention_k=16, attention_heads=4, attention_dropout=0.1, pt_k=16)
N_POINTS = 4096
WARMUP, REPS = 5, 20


def load_model(path):
    ck = torch.load(path, map_location="cpu", weights_only=False)
    meta = {**DEFAULTS}
    for key in DEFAULTS:
        if ck.get(key) is not None:
            meta[key] = ck[key]
    mtype = ck.get("model_type", "pointnet")
    m = build_model(
        mtype, task="segmentation",
        num_classes=ck["num_classes"], num_features=ck["num_features"],
        k=meta["k"], k_small=meta["k_small"], k_large=meta["k_large"],
        attention_type=meta["attention_type"], attention_k=meta["attention_k"],
        attention_heads=meta["attention_heads"], attention_dropout=meta["attention_dropout"],
        pt_k=meta["pt_k"],
    )
    m.load_state_dict(ck["model_state_dict"], strict=False)
    return m.eval(), ck["num_features"]


def get_macs_torchinfo_grouped(model, num_features):
    """
    MACs по каждому named_child:
    - Запускаем полный forward, захватывая входной тензор каждого блока.
    - Затем прогоняем thop отдельно на каждом блоке с захваченным входом.
    """
    from thop import profile as thop_profile
    captured = {}
    handles = []

    for name, module in model.named_children():
        def make_hook(n):
            def hook(mod, inp):
                if inp and isinstance(inp[0], torch.Tensor):
                    captured[n] = inp[0].detach().cpu()
            return hook
        handles.append(module.register_forward_pre_hook(make_hook(name)))

    x = torch.randn(1, N_POINTS, num_features)
    with torch.no_grad():
        try:
            model(x)
        except Exception:
            pass
    for h in handles:
        h.remove()

    grouped = {}
    for name, module in model.named_children():
        inp = captured.get(name)
        if inp is None:
            grouped[name] = 0
            continue
        try:
            macs, _ = thop_profile(module, inputs=(inp,), verbose=False)
            grouped[name] = macs
        except Exception:
            grouped[name] = 0
    return grouped


def get_timing_per_layer(model, num_features, device):
    """Реальное время (мс) по каждому named_children через CUDA events."""
    model = model.to(device)
    x = torch.randn(1, N_POINTS, num_features, device=device)
    timings = {n: [] for n, _ in model.named_children()}
    handles = []

    for name, module in model.named_children():
        s = torch.cuda.Event(enable_timing=True)
        e = torch.cuda.Event(enable_timing=True)

        def make(n, s, e):
            def pre(m, i): s.record()
            def post(m, i, o):
                e.record()
                torch.cuda.synchronize()
                timings[n].append(s.elapsed_time(e))
            return pre, post

        pre, post = make(name, s, e)
        handles.append(module.register_forward_pre_hook(pre))
        handles.append(module.register_forward_hook(post))

    with torch.no_grad():
        for _ in range(WARMUP):
            model(x)
    timings = {n: [] for n in timings}
    with torch.no_grad():
        for _ in range(REPS):
            model(x)
    for h in handles:
        h.remove()

    avg = {k: float(np.mean(v)) for k, v in timings.items() if v}
    model.cpu()
    torch.cuda.empty_cache()
    return avg


# ── Сбор данных ───────────────────────────────────────────────────────────────

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
all_data = []  # list of (model_label, block_name, gmacs, ms, pct_macs, pct_time)

for label, mtype, path in CONFIGS:
    print(f"  {label}...")
    # Отдельные экземпляры модели для MACs (CPU) и timing (GPU)
    model_for_macs, num_feat = load_model(path)
    model_for_time, _        = load_model(path)

    blocks = [n for n, _ in model_for_macs.named_children()]

    macs_map = get_macs_torchinfo_grouped(model_for_macs, num_feat)
    del model_for_macs

    time_map = get_timing_per_layer(model_for_time, num_feat, device)
    del model_for_time

    total_macs = sum(macs_map.values()) or 1
    total_time = sum(time_map.values()) or 1
    for block in blocks:
        mmacs = macs_map.get(block, 0) / 1e6
        ms    = time_map.get(block, 0)
        pct_macs = macs_map.get(block, 0) / total_macs * 100
        pct_time = ms / total_time * 100
        all_data.append((label, block, mmacs, ms, pct_macs, pct_time))

    # Итого строка
    all_data.append((label, "ИТОГО", total_macs / 1e6, total_time, 100.0, 100.0))
    print(f"    total MACs={total_macs/1e6:.0f}M  total_time={total_time:.1f}ms")

# ── Excel ─────────────────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "MACs + Время по слоям"

HDR   = PatternFill("solid", start_color="BDD7EE")
TOTAL = PatternFill("solid", start_color="D9D9D9")
OR    = PatternFill("solid", start_color="F4B942")
YE    = PatternFill("solid", start_color="FFEB9C")
MODEL_COLORS = ["DEEAF1","E2EFDA","FFF2CC","FCE4D6","EAE1F4","D9EAD3","FDE9D9","E8F0FE"]
thin = Side(style="thin", color="AAAAAA")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

model_list = [c[0] for c in CONFIGS]
mc = {m: MODEL_COLORS[i] for i, m in enumerate(model_list)}

headers = ["Модель", "Блок", "MACs (млн)", "% от MACs", "Время (мс)", "% от времени"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col)
    c.font = Font(bold=True, name="Arial", size=10)
    c.fill = HDR
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd
ws.freeze_panes = "A2"

for r_idx, (model, block, mmacs, ms, pct_m, pct_t) in enumerate(all_data, 2):
    is_total = block == "ИТОГО"
    fill = TOTAL if is_total else PatternFill("solid", start_color=mc[model])
    vals = [model, block,
            round(mmacs, 1) if mmacs else 0,
            round(pct_m, 1) if pct_m else 0,
            round(ms, 3) if ms else 0,
            round(pct_t, 1) if pct_t else 0]
    for col, val in enumerate(vals, 1):
        c = ws.cell(r_idx, col)
        c.value = val
        c.font = Font(bold=is_total, name="Arial", size=10)
        c.fill = fill
        c.border = brd
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if col in (3, 4, 5, 6) and not is_total:
            c.number_format = "0.0"

    # Подсветка % времени
    if not is_total and pct_t is not None:
        ct = ws.cell(r_idx, 6)
        if pct_t >= 50:
            ct.fill = OR
            ct.font = Font(bold=True, name="Arial", size=10)
        elif pct_t >= 15:
            ct.fill = YE

    # Подсветка % MACs
    if not is_total and pct_m is not None:
        cm = ws.cell(r_idx, 4)
        if pct_m >= 50:
            cm.fill = OR
            cm.font = Font(bold=True, name="Arial", size=10)
        elif pct_m >= 15:
            cm.fill = YE

ws.column_dimensions["A"].width = 17
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 14
for r in range(2, len(all_data) + 2):
    ws.row_dimensions[r].height = 22

# ── Лист 2: сводная таблица по моделям ────────────────────────────────────────
ws2 = wb.create_sheet("Сводка по моделям")
ws2.append(["Модель", "Всего MACs (млн)", "Время инференса (мс)", "MACs (млн)/мс"])
for c in range(1, 5):
    ws2.cell(1, c).font = Font(bold=True, name="Arial")
    ws2.cell(1, c).fill = HDR
    ws2.cell(1, c).border = brd
    ws2.cell(1, c).alignment = Alignment(horizontal="center")

totals_by_model = {}
for model, block, mmacs, ms, pct_m, pct_t in all_data:
    if block == "ИТОГО":
        totals_by_model[model] = (mmacs, ms)

for i, label in enumerate(model_list, 2):
    if label in totals_by_model:
        mmacs, ms = totals_by_model[label]
        fill = PatternFill("solid", start_color=mc[label])
        eff = round(mmacs / ms, 1) if ms > 0 else 0
        for col, val in enumerate([label, round(mmacs, 0), round(ms,1), eff], 1):
            c = ws2.cell(i, col)
            c.value = val
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.border = brd
            c.alignment = Alignment(horizontal="center")

for col, w in zip("ABCD", [18, 14, 22, 20]):
    ws2.column_dimensions[col].width = w

# График MACs
chart_m = BarChart()
chart_m.type = "col"
chart_m.title = "Вычислительная сложность (MACs, млн)"
chart_m.y_axis.title = "MACs (млн)"
chart_m.style = 10
chart_m.width = 18
chart_m.height = 12
data_m = Reference(ws2, min_col=2, min_row=1, max_row=len(model_list)+1)
cats   = Reference(ws2, min_col=1, min_row=2, max_row=len(model_list)+1)
chart_m.add_data(data_m, titles_from_data=True)
chart_m.set_categories(cats)
ws2.add_chart(chart_m, "F2")

# График времени
chart_t = BarChart()
chart_t.type = "col"
chart_t.title = "Время инференса (мс)"
chart_t.y_axis.title = "мс"
chart_t.style = 10
chart_t.width = 18
chart_t.height = 12
data_t = Reference(ws2, min_col=3, min_row=1, max_row=len(model_list)+1)
chart_t.add_data(data_t, titles_from_data=True)
chart_t.set_categories(cats)
ws2.add_chart(chart_t, "F20")

out = "results/combined_analysis.xlsx"
wb.save(out)
print("saved:", out)
