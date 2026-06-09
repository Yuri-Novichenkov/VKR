"""
Анализ чувствительности к числу точек N (num_points).

Для каждой модели и каждого N измеряет:
  1. Скорость инференса (мс, CPU или GPU через --device)
  2. Точность (mIoU) на тестовом наборе Mar16 — с scene-level voting

Модели: все 8 — PointNet, PointNet++, DGCNN, LDGCNN, LDGCNN-GATv2,
        LDGCNN-LocalWindow, LDGCNNFlash, PointTransformer
Значения N: 512, 1024, 2048, 3072, 4096 (baseline)

Запуск (ноутбук, CPU):
    # только скорость (~5-10 мин):
    python scripts/n_sensitivity.py --speed_only --device cpu

    # скорость + точность (~2-3 часа):
    python scripts/n_sensitivity.py --device cpu

    # только конкретные модели:
    python scripts/n_sensitivity.py --speed_only --device cpu --models PointNet DGCNN

Запуск (сервер, GPU):
    python scripts/n_sensitivity.py --speed_only --device cuda
    python scripts/n_sensitivity.py --device cuda
"""

import sys, argparse, time, torch, numpy as np
from pathlib import Path

sys.path.insert(0, ".")
from src.models import build_model
from src.utils.metrics import calculate_metrics

# ── Чекпоинты лучших конфигураций на Mar16 ───────────────────────────────────

CONFIGS = [
    ("PointNet",         "pointnet",         "checkpoints/loss_sweep/pointnet/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointNet++",       "pointnet++",       "checkpoints/loss_sweep/pointnet++/segmentation/loss_lovasz_g2p0__cb_effective_b0p99999/mar16/best_model.pth"),
    ("DGCNN",            "dgcnn",            "checkpoints/loss_sweep/dgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN",           "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN-GATv2",     "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_gatv2_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNN-LocalWin",  "ldgcnn",           "checkpoints/loss_sweep/ldgcnn/segmentation/attn_local_window_k16_h4_d0p1__cb_effective_b0p99999/mar16/best_model.pth"),
    ("LDGCNNFlash",      "ldgcnn_flash",     "checkpoints/loss_sweep/ldgcnn_flash/segmentation/cb_effective_b0p99999/mar16/best_model.pth"),
    ("PointTransformer", "pointtransformer", "checkpoints/pointtransformer/segmentation/cb_effective_b0p999/mar16/best_model.pth"),
]

DEFAULTS = dict(k=20, k_small=20, k_large=40, attention_type="none",
                attention_k=16, attention_heads=4, attention_dropout=0.1, pt_k=16)

N_VALUES  = [512, 1024, 2048, 3072, 4096]
N_BASELINE = 4096
TEST_DATA  = "Files/Mar16/LiDAR/Mar16_test_GroundTruth.laz"
WARMUP, REPS = 5, 20


# ── Загрузка модели ───────────────────────────────────────────────────────────

def load_model(path):
    ck = torch.load(path, map_location="cpu", weights_only=False)
    meta = {**DEFAULTS}
    for key in DEFAULTS:
        if ck.get(key) is not None:
            meta[key] = ck[key]
    mtype = ck.get("model_type", "pointnet")
    m = build_model(
        mtype, task="segmentation",
        num_classes=ck["num_classes"], num_features=ck["num_features"],
        k=meta["k"], k_small=meta["k_small"], k_large=meta["k_large"],
        attention_type=meta["attention_type"], attention_k=meta["attention_k"],
        attention_heads=meta["attention_heads"], attention_dropout=meta["attention_dropout"],
        pt_k=meta["pt_k"],
    )
    m.load_state_dict(ck["model_state_dict"], strict=False)
    return m.eval(), ck["num_features"], ck["num_classes"], mtype


# ── Бенчмарк скорости ─────────────────────────────────────────────────────────

def bench_speed(model, num_features, n_points, device):
    model = model.to(device)
    x = torch.randn(1, n_points, num_features, device=device)

    use_cuda = device.type == "cuda"

    def _run(x):
        out = model(x)
        return out[0] if isinstance(out, tuple) else out

    with torch.no_grad():
        for _ in range(WARMUP):
            _run(x)
    if use_cuda:
        torch.cuda.synchronize()

    times = []
    with torch.no_grad():
        if use_cuda:
            s = torch.cuda.Event(enable_timing=True)
            e = torch.cuda.Event(enable_timing=True)
            for _ in range(REPS):
                s.record(); _run(x); e.record()
                torch.cuda.synchronize()
                times.append(s.elapsed_time(e))
        else:
            for _ in range(REPS):
                t0 = time.perf_counter()
                _run(x)
                times.append((time.perf_counter() - t0) * 1000)

    model.cpu()
    if use_cuda:
        torch.cuda.empty_cache()
    return float(np.mean(times)), float(np.std(times))


# ── Оценка точности (mIoU с voting) ──────────────────────────────────────────

def make_test_dataset(n_points):
    """Создаёт датасет с заданным N. Кэш индивидуален для каждого N."""
    from src.data.dataset import LiDARDataset
    return LiDARDataset(
        TEST_DATA, num_points=n_points, augment=False,
        task="segmentation",
        cache_dir="cache", cache_mode="write",
    )


def eval_accuracy(model, num_classes, device, dataset):
    from torch.utils.data import DataLoader
    loader = DataLoader(dataset, batch_size=4, shuffle=False,
                        num_workers=0, pin_memory=(device.type == "cuda"))

    num_total = len(dataset.features)
    vote_counts = np.zeros((num_total, num_classes), dtype=np.int32)
    model = model.to(device)

    with torch.no_grad():
        sample_offset = 0
        for features, _labels in loader:
            features = features.to(device)
            out = model(features)
            logits = out[0] if isinstance(out, tuple) else out  # PointNet возвращает (x, t1, t2)
            preds = logits.argmax(dim=-1).cpu().numpy()
            bsz = preds.shape[0]
            for b in range(bsz):
                idx = dataset.get_cloud_point_indices(sample_offset + b)
                np.add.at(vote_counts, (idx, preds[b]), 1)
            sample_offset += bsz

    aggregated = vote_counts.argmax(axis=1)
    true_labels = np.asarray(dataset.labels, dtype=np.int64)
    metrics = calculate_metrics(
        torch.from_numpy(aggregated),
        torch.from_numpy(true_labels),
        num_classes=num_classes,
    )
    model.cpu()
    if device.type == "cuda":
        torch.cuda.empty_cache()
    return float(metrics["mean_iou"]) * 100


# ── Основной цикл ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--speed_only", action="store_true",
                        help="Только скорость, без оценки точности")
    parser.add_argument("--device", default="auto", choices=["auto", "cpu", "cuda"],
                        help="Устройство (default: auto)")
    parser.add_argument("--models", nargs="+", default=None,
                        help="Запустить только для указанных моделей, напр. PointNet DGCNN")
    parser.add_argument("--n_values", nargs="+", type=int, default=N_VALUES,
                        help=f"Значения N для тестирования (default: {N_VALUES})")
    args = parser.parse_args()

    if args.device == "cpu":
        device = torch.device("cpu")
    elif args.device == "cuda":
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    else:
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    print(f"Устройство: {device}")
    if device.type == "cuda":
        print(f"GPU: {torch.cuda.get_device_name(device)}")

    test_exists = Path(TEST_DATA).exists()
    if not args.speed_only and not test_exists:
        print(f"WARN: тестовый файл не найден ({TEST_DATA}), только скорость.")
        args.speed_only = True

    # Датасеты создаются один раз на N, переиспользуются по всем моделям.
    # num_workers=0 на Windows чтобы избежать pickle-проблем с Dataset.
    datasets = {}
    if not args.speed_only:
        print("\nЗагрузка тестовых датасетов для каждого N (первый раз — с диска)...")
        for n in args.n_values:
            print(f"  N={n}...", end=" ", flush=True)
            datasets[n] = make_test_dataset(n)
            print(f"{len(datasets[n])} окон")

    all_results = []

    for label, mtype, path in CONFIGS:
        if args.models and label not in args.models:
            continue
        if not Path(path).exists():
            print(f"\nWARN: чекпоинт не найден, пропускаю: {path}")
            continue

        print(f"\n{'='*58}")
        print(f"  {label}")
        print(f"{'='*58}")

        model, num_feat, num_cls, _ = load_model(path)

        # baseline для расчёта speedup
        baseline_ms, _ = bench_speed(model, num_feat, N_BASELINE, device)
        print(f"  baseline N={N_BASELINE}: {baseline_ms:.1f} мс")
        print(f"  {'N':>5}  {'мс':>8}  {'±мс':>6}  {'speedup':>8}  {'mIoU%':>8}")
        print(f"  {'-'*42}")

        for n in args.n_values:
            ms, std = bench_speed(model, num_feat, n, device)

            miou = None
            if not args.speed_only and n in datasets:
                miou = eval_accuracy(model, num_cls, device, datasets[n])

            speedup = baseline_ms / ms
            marker = " <- baseline" if n == N_BASELINE else ""
            miou_str = f"{miou:.2f}" if miou is not None else "—"
            print(f"  {n:>5}  {ms:>8.1f}  {std:>6.1f}  {speedup:>8.2f}x  {miou_str:>8}{marker}")

            all_results.append({
                "model":      label,
                "N":          n,
                "baseline_N": N_BASELINE,
                "device":     str(device),
                "ms":         round(ms, 1),
                "std_ms":     round(std, 1),
                "speedup":    round(speedup, 2),
                "miou":       round(miou, 2) if miou is not None else None,
            })

    # ── Консольная сводка ──────────────────────────────────────────────────────
    print("\n\n" + "="*70)
    print("ИТОГОВАЯ ТАБЛИЦА — чувствительность к N")
    print("="*70)
    cur = None
    for r in all_results:
        if r["model"] != cur:
            cur = r["model"]
            print(f"\n  {cur} (baseline N={r['baseline_N']}, device={r['device']})")
            print(f"  {'N':>5}  {'мс':>8}  {'speedup':>8}  {'mIoU%':>8}")
            print(f"  {'-'*34}")
        marker = " *base*" if r["N"] == r["baseline_N"] else ""
        miou_s = f"{r['miou']:.2f}" if r["miou"] is not None else "—"
        print(f"  {r['N']:>5}  {r['ms']:>8.1f}  {r['speedup']:>8.2f}x  {miou_s:>8}{marker}")

    # ── Excel ──────────────────────────────────────────────────────────────────
    _save_excel(all_results, args.device if args.device != "auto" else str(device.type))
    print(f"\nГотово.")


def _save_excel(all_results, device_tag):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "N sensitivity"

    HDR  = PatternFill("solid", start_color="BDD7EE")
    BASE = PatternFill("solid", start_color="C6EFCE")
    thin = Side(style="thin", color="AAAAAA")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    PALETTE = ["DEEAF1", "E2EFDA", "FFF2CC", "FCE4D6", "EAE1F4", "D9EAD3", "FDE9D9", "E8F5E9"]
    model_names = list(dict.fromkeys(r["model"] for r in all_results))
    mc = {m: PALETTE[i % len(PALETTE)] for i, m in enumerate(model_names)}

    # ── Лист 1: таблица данных ─────────────────────────────────────────────────
    headers = ["Модель", "N (точек)", "Baseline N", "Устройство",
               "Время (мс)", "Std (мс)", "Ускорение (x)", "mIoU (%)"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col)
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = brd

    for r_idx, r in enumerate(all_results, 2):
        is_base = r["N"] == r["baseline_N"]
        fill = BASE if is_base else PatternFill("solid", start_color=mc[r["model"]])
        vals = [r["model"], r["N"], r["baseline_N"], r["device"],
                r["ms"], r["std_ms"], r["speedup"],
                r["miou"] if r["miou"] is not None else "—"]
        for col, val in enumerate(vals, 1):
            c = ws.cell(r_idx, col)
            c.value = val
            c.font = Font(bold=is_base, name="Arial", size=10)
            c.fill = fill
            c.border = brd
            c.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFGH", [16, 10, 10, 10, 12, 10, 14, 10]):
        ws.column_dimensions[col].width = w

    # ── Лист 2: данные для графиков ───────────────────────────────────────────
    ws2 = wb.create_sheet("Графики")
    n_all = sorted(set(r["N"] for r in all_results))

    # Скорость (мс)
    ws2.cell(1, 1).value = "N"
    col_map = {}
    for i, m in enumerate(model_names, 2):
        ws2.cell(1, i).value = m
        col_map[m] = i
    for row_i, n in enumerate(n_all, 2):
        ws2.cell(row_i, 1).value = n
        for m in model_names:
            vals = [r["ms"] for r in all_results if r["model"] == m and r["N"] == n]
            ws2.cell(row_i, col_map[m]).value = vals[0] if vals else None

    chart_ms = LineChart()
    chart_ms.title = "Время инференса vs N точек"
    chart_ms.y_axis.title = "мс"
    chart_ms.x_axis.title = "N (число точек)"
    chart_ms.style = 10
    chart_ms.width = 24
    chart_ms.height = 14
    for m in model_names:
        data = Reference(ws2, min_col=col_map[m], min_row=1, max_row=len(n_all) + 1)
        chart_ms.add_data(data, titles_from_data=True)
    chart_ms.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=len(n_all) + 1))
    ws2.add_chart(chart_ms, "A" + str(len(n_all) + 4))

    # Ускорение (x)
    sp_row_start = len(n_all) + 3
    ws2.cell(sp_row_start, 1).value = "N"
    sp_col = {}
    for i, m in enumerate(model_names, 2):
        ws2.cell(sp_row_start, i).value = f"{m} speedup"
        sp_col[m] = i
    for row_i, n in enumerate(n_all, sp_row_start + 1):
        ws2.cell(row_i, 1).value = n
        for m in model_names:
            vals = [r["speedup"] for r in all_results if r["model"] == m and r["N"] == n]
            ws2.cell(row_i, sp_col[m]).value = vals[0] if vals else None

    chart_sp = LineChart()
    chart_sp.title = "Ускорение vs N точек (относительно N=4096)"
    chart_sp.y_axis.title = "Ускорение (x)"
    chart_sp.x_axis.title = "N"
    chart_sp.style = 10
    chart_sp.width = 24
    chart_sp.height = 14
    for m in model_names:
        data = Reference(ws2, min_col=sp_col[m], min_row=sp_row_start,
                         max_row=sp_row_start + len(n_all))
        chart_sp.add_data(data, titles_from_data=True)
    chart_sp.set_categories(Reference(ws2, min_col=1,
                                       min_row=sp_row_start + 1,
                                       max_row=sp_row_start + len(n_all)))
    ws2.add_chart(chart_sp, "M" + str(len(n_all) + 4))

    # mIoU (если есть данные)
    if any(r["miou"] is not None for r in all_results):
        miou_row_start = sp_row_start + len(n_all) + 3
        ws2.cell(miou_row_start, 1).value = "N"
        mi_col = {}
        for i, m in enumerate(model_names, 2):
            ws2.cell(miou_row_start, i).value = f"{m} mIoU"
            mi_col[m] = i
        for row_i, n in enumerate(n_all, miou_row_start + 1):
            ws2.cell(row_i, 1).value = n
            for m in model_names:
                vals = [r["miou"] for r in all_results
                        if r["model"] == m and r["N"] == n and r["miou"] is not None]
                ws2.cell(row_i, mi_col[m]).value = vals[0] if vals else None

        chart_miou = LineChart()
        chart_miou.title = "mIoU vs N точек"
        chart_miou.y_axis.title = "mIoU (%)"
        chart_miou.x_axis.title = "N"
        chart_miou.style = 10
        chart_miou.width = 24
        chart_miou.height = 14
        for m in model_names:
            data = Reference(ws2, min_col=mi_col[m], min_row=miou_row_start,
                             max_row=miou_row_start + len(n_all))
            chart_miou.add_data(data, titles_from_data=True)
        chart_miou.set_categories(Reference(ws2, min_col=1,
                                             min_row=miou_row_start + 1,
                                             max_row=miou_row_start + len(n_all)))
        ws2.add_chart(chart_miou, "A" + str(miou_row_start + len(n_all) + 3))

    out = f"results/n_sensitivity_{device_tag}.xlsx"
    Path("results").mkdir(exist_ok=True)
    wb.save(out)
    print(f"Сохранено: {out}")


if __name__ == "__main__":
    main()
